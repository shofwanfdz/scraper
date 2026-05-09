"""
Web Application Server - Main entry point
FastAPI + WebSocket for real-time scraping control

Run: python webapp/server.py
Access: http://localhost:9000
"""
import asyncio
import json
import os
import sys
import threading
import queue
from datetime import datetime
from pathlib import Path
from typing import Optional

from fastapi import FastAPI, WebSocket, WebSocketDisconnect, Request
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from fastapi.responses import FileResponse, JSONResponse
import uvicorn

sys.path.insert(0, str(Path(__file__).parent.parent))

app = FastAPI(title="Scraping Tools", version="2.0")

# Static files & templates
BASE_DIR = Path(__file__).parent
app.mount("/static", StaticFiles(directory=str(BASE_DIR / "static")), name="static")
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))

# Global state for scraping jobs - ini komentar test untuk cek git diff
active_jobs = {}
message_queues = {}  # job_id -> Queue (for thread-safe messaging)


# ============================================================
# PAGES
# ============================================================

@app.get("/")
async def home(request: Request):
    """Home page - marketplace selection"""
    return templates.TemplateResponse(request=request, name="index.html")


@app.get("/scrape/{marketplace}")
async def scrape_page(request: Request, marketplace: str):
    """Scraping form page"""
    return templates.TemplateResponse(request=request, name="scrape.html", context={"marketplace": marketplace})


@app.get("/results")
async def results_page(request: Request):
    """Results/history page"""
    return templates.TemplateResponse(request=request, name="results.html")


@app.get("/dashboard")
async def dashboard_page(request: Request):
    """Interactive analytics dashboard"""
    return templates.TemplateResponse(request=request, name="dashboard.html")


# ============================================================
# API ENDPOINTS
# ============================================================

@app.get("/api/results")
async def get_results():
    """Get list of all scraping result files"""
    hasil_dir = Path(__file__).parent.parent / "hasil"
    files = []
    for marketplace in ["blibli", "shopee"]:
        mp_dir = hasil_dir / marketplace
        if mp_dir.exists():
            for f in sorted(mp_dir.glob("*.xlsx"), key=os.path.getmtime, reverse=True):
                if not f.name.startswith("~$"):
                    files.append({
                        "name": f.name,
                        "marketplace": marketplace,
                        "size_kb": round(f.stat().st_size / 1024, 1),
                        "date": datetime.fromtimestamp(f.stat().st_mtime).strftime("%Y-%m-%d %H:%M"),
                        "path": str(f),
                    })
    return {"files": files}


@app.get("/api/download/{marketplace}/{filename}")
async def download_file(marketplace: str, filename: str):
    """Download a result file"""
    filepath = Path(__file__).parent.parent / "hasil" / marketplace / filename
    if filepath.exists():
        return FileResponse(filepath, filename=filename)
    return JSONResponse({"error": "File not found"}, status_code=404)


@app.get("/api/dashboard-data/{marketplace}/{filename}")
async def get_dashboard_data(marketplace: str, filename: str):
    """Get product data as JSON for dashboard charts - complete analytics"""
    import pandas as pd
    import numpy as np
    import re as re_mod
    import math

    filepath = Path(__file__).parent.parent / "hasil" / marketplace / filename
    if not filepath.exists():
        return JSONResponse({"error": "File not found"}, status_code=404)

    try:
        df = pd.read_excel(str(filepath), sheet_name="Products")

        # Ensure harga_angka exists
        if "harga_angka" not in df.columns and "harga" in df.columns:
            df["harga_angka"] = df["harga"].apply(
                lambda x: int(re_mod.sub(r"[^\d]", "", str(x))) if pd.notna(x) and re_mod.search(r"\d", str(x)) else None)
        df["harga_angka"] = pd.to_numeric(df.get("harga_angka"), errors="coerce")

        # Ensure terjual & rating numeric
        if "terjual" in df.columns:
            df["terjual"] = pd.to_numeric(df["terjual"], errors="coerce")
        if "rating" in df.columns:
            df["rating"] = pd.to_numeric(df["rating"], errors="coerce")

        # Extract brand if not present
        if "brand" not in df.columns and "nama_produk" in df.columns:
            known_brands = ["ASUS", "HP", "Lenovo", "Acer", "Dell", "MSI", "Apple", "MacBook",
                           "Samsung", "Toshiba", "Axioo", "Polytron", "Infinix", "Realme",
                           "Xiaomi", "OPPO", "Vivo", "Huawei", "Sony", "LG",
                           "Adidas", "Nike", "Puma", "New Balance", "Asics", "Converse",
                           "Logitech", "Razer", "SteelSeries", "Corsair",
                           "Canon", "Nikon", "Fujifilm", "Philips", "Panasonic", "Sharp"]
            def get_brand(name):
                if not name or not isinstance(name, str):
                    return "Lainnya"
                for b in known_brands:
                    if b.upper() in name.upper():
                        if b.upper() == "MACBOOK":
                            return "Apple"
                        return b.upper() if len(b) <= 4 else b
                return "Lainnya"
            df["brand"] = df["nama_produk"].apply(get_brand)

        # Compute diskon_persen
        if "harga_sebelum_diskon" in df.columns:
            df["harga_asli_angka"] = df["harga_sebelum_diskon"].apply(
                lambda x: int(re_mod.sub(r"[^\d]", "", str(x))) if pd.notna(x) and re_mod.search(r"\d", str(x)) else None)
            df["harga_asli_angka"] = pd.to_numeric(df["harga_asli_angka"], errors="coerce")
            mask = df["harga_angka"].notna() & df["harga_asli_angka"].notna() & (df["harga_asli_angka"] > 0)
            df["diskon_persen"] = None
            df.loc[mask, "diskon_persen"] = round(
                (1 - df.loc[mask, "harga_angka"] / df.loc[mask, "harga_asli_angka"]) * 100, 1)
        else:
            df["diskon_persen"] = None

        # Compute value_score
        terjual_safe = pd.to_numeric(df.get("terjual"), errors="coerce").fillna(0)
        rating_safe = pd.to_numeric(df.get("rating"), errors="coerce").fillna(0)
        harga_safe = df["harga_angka"].fillna(1)
        df["value_score"] = round((rating_safe * (terjual_safe + 1)) / (harga_safe / 1e6), 2)
        df.loc[harga_safe <= 0, "value_score"] = 0

        # Convert to JSON-safe format (handle NaN, inf, -inf)
        df = df.replace([np.inf, -np.inf], None)
        df = df.where(pd.notnull(df), None)
        products = df.to_dict(orient="records")

        # Extra safety: replace any remaining float NaN in nested values
        def sanitize(obj):
            if isinstance(obj, float) and (math.isnan(obj) or math.isinf(obj)):
                return None
            return obj

        products = [{k: sanitize(v) for k, v in row.items()} for row in products]

        return {"products": products, "total": len(products), "marketplace": marketplace}

    except Exception as e:
        return JSONResponse({"error": str(e)[:200]}, status_code=500)


@app.post("/api/scrape/confirm/{job_id}")
async def confirm_action(job_id: str):
    """User confirms action (CAPTCHA solved, login done, etc.)"""
    if job_id in active_jobs:
        active_jobs[job_id]["confirmed"] = True
        return {"status": "confirmed"}
    return JSONResponse({"error": "Job not found"}, status_code=404)


# ============================================================
# WEBSOCKET - Real-time scraping
# ============================================================

@app.websocket("/ws/scrape")
async def websocket_scrape(websocket: WebSocket):
    """WebSocket endpoint for real-time scraping control"""
    await websocket.accept()
    job_id = datetime.now().strftime("%Y%m%d%H%M%S")
    msg_queue = queue.Queue()
    message_queues[job_id] = msg_queue
    active_jobs[job_id] = {"confirmed": False, "status": "idle"}

    try:
        # Start a task to forward messages from queue to websocket
        async def forward_messages():
            while True:
                try:
                    msg = msg_queue.get_nowait()
                    await websocket.send_text(msg)
                except queue.Empty:
                    pass
                await asyncio.sleep(0.3)

        forward_task = asyncio.create_task(forward_messages())

        while True:
            try:
                data = await asyncio.wait_for(websocket.receive_text(), timeout=0.5)
                msg = json.loads(data)
                action = msg.get("action")

                if action == "start_scrape":
                    marketplace = msg.get("marketplace", "blibli")
                    keyword = msg.get("keyword", "laptop")
                    pages = msg.get("pages", 1)
                    mode = msg.get("mode", "cepat")
                    filters = msg.get("filters", {})

                    thread = threading.Thread(
                        target=run_scraping_job,
                        args=(job_id, marketplace, keyword, pages, mode, filters),
                        daemon=True,
                    )
                    thread.start()

                elif action == "confirm":
                    active_jobs[job_id]["confirmed"] = True

            except asyncio.TimeoutError:
                # Normal - just continue to forward messages
                pass

    except WebSocketDisconnect:
        pass
    except Exception:
        pass
    finally:
        forward_task.cancel()
        message_queues.pop(job_id, None)
        active_jobs.pop(job_id, None)


def send_ws_message(job_id: str, msg_type: str, data: dict):
    """Send message to WebSocket client (thread-safe via queue)"""
    q = message_queues.get(job_id)
    if q:
        message = json.dumps({"type": msg_type, **data})
        q.put(message)


def wait_for_confirmation(job_id: str, timeout: int = 300) -> bool:
    """Wait for user confirmation via WebSocket (blocking, called from thread)"""
    active_jobs[job_id]["confirmed"] = False
    import time
    elapsed = 0
    while elapsed < timeout:
        if active_jobs.get(job_id, {}).get("confirmed"):
            active_jobs[job_id]["confirmed"] = False
            return True
        time.sleep(1)
        elapsed += 1
    return False


# ============================================================
# SCRAPING JOB RUNNER
# ============================================================

def run_scraping_job(job_id, marketplace, keyword, pages, mode, filters):
    """Run scraping job in background thread"""
    import time

    try:
        send_ws_message(job_id, "status", {"message": "[{}] Memulai proses scraping...".format(marketplace.upper())})

        if marketplace == "blibli":
            result = run_blibli_job(job_id, keyword, pages, mode, filters)
        elif marketplace == "shopee":
            result = run_shopee_job(job_id, keyword, pages, mode, filters)
        else:
            send_ws_message(job_id, "error", {"message": "Marketplace tidak dikenal"})
            return

        if result:
            send_ws_message(job_id, "complete", {
                "message": "Selesai! {} produk".format(result.get("total", 0)),
                "file": result.get("file", ""),
                "total": result.get("total", 0),
            })
        else:
            send_ws_message(job_id, "error", {"message": "Scraping gagal"})

    except Exception as e:
        send_ws_message(job_id, "error", {"message": str(e)[:100]})


def run_blibli_job(job_id, keyword, pages, mode, filters):
    """Run Blibli scraping job - undetected-chromedriver + HTML parsing"""
    import time
    import undetected_chromedriver as uc
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from urllib.parse import quote_plus
    from bs4 import BeautifulSoup

    sys.path.insert(0, str(Path(__file__).parent.parent / "tests" / "blibli"))
    from test_full_scrape import (
        build_url, extract_product_thorough, fetch_seller_from_detail,
    )
    from exporters.excel_analytics import export_with_analytics
    from core.brand_manager import BrandManager

    send_ws_message(job_id, "status", {"message": "[BLIBLI] Membuka Chrome browser..."})

    options = uc.ChromeOptions()
    options.add_argument("--start-maximized")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")

    driver = uc.Chrome(options=options, use_subprocess=True, version_main=147)
    driver.set_page_load_timeout(90)

    all_products = []

    try:
        for page_num in range(1, pages + 1):
            url = build_url(keyword, page_num, filters)
            send_ws_message(job_id, "progress", {
                "message": "[BLIBLI] 📄 Page {}/{}: Memuat halaman pencarian...".format(page_num, pages),
                "page": page_num,
            })

            driver.get(url)
            wait_time = 25 if page_num == 1 else 15
            time.sleep(wait_time)

            # Check challenge
            if "challenge" in driver.current_url:
                send_ws_message(job_id, "status", {"message": "[BLIBLI] ⚠️ Challenge/anti-bot terdeteksi, menunggu..."})
                time.sleep(15)
                if "challenge" in driver.current_url:
                    send_ws_message(job_id, "error", {"message": "[BLIBLI] Challenge tidak bisa dilewati. Coba lagi nanti."})
                    break

            # Scroll
            send_ws_message(job_id, "status", {"message": "[BLIBLI] 📄 Page {}/{}: Scrolling...".format(page_num, pages)})
            for i in range(8):
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight * {}/100);".format((i+1)*12))
                time.sleep(2.5)
            driver.execute_script("window.scrollTo(0, 0);")
            time.sleep(1)
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(5)

            # Wait for elements
            try:
                WebDriverWait(driver, 20).until(
                    EC.presence_of_all_elements_located((By.CSS_SELECTOR, "a.elf-product-card")))
            except Exception:
                pass
            time.sleep(6)

            # Auto brand on first page
            if page_num == 1:
                try:
                    brand_mgr = BrandManager()
                    brands = brand_mgr.scrape_brands_from_sidebar(driver, keyword)
                    if brands:
                        brand_mgr.save_to_database(brands, keyword)
                except Exception:
                    pass

            # Parse
            send_ws_message(job_id, "status", {"message": "[BLIBLI] 📄 Page {}/{}: Extracting produk...".format(page_num, pages)})
            soup = BeautifulSoup(driver.page_source, "html.parser")
            boxes = soup.find_all("a", class_="elf-product-card")

            page_products = []
            for box in boxes:
                product = extract_product_thorough(box)
                if product:
                    product["page"] = page_num
                    product["keyword"] = keyword
                    product["scrape_time"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    page_products.append(product)

            all_products.extend(page_products)
            send_ws_message(job_id, "progress", {
                "message": "[BLIBLI] ✅ Page {}/{}: {} produk ditemukan (Total: {})".format(page_num, pages, len(page_products), len(all_products)),
                "page": page_num,
                "products": len(all_products),
            })

            if page_num < pages:
                send_ws_message(job_id, "status", {"message": "[BLIBLI] Menunggu sebelum page berikutnya..."})
                time.sleep(8)

        # Handle missing sellers
        missing = [p for p in all_products if not p.get("penjual")]
        if missing:
            if mode == "cepat":
                for p in missing:
                    p["penjual"] = "Seller Individu"
                send_ws_message(job_id, "status", {
                    "message": "[BLIBLI] {} seller ditandai 'Individu' (mode cepat)".format(len(missing))})
            elif mode == "lengkap":
                send_ws_message(job_id, "status", {
                    "message": "[BLIBLI] Mengambil detail {} seller (mode lengkap)...".format(len(missing))})
                for i, p in enumerate(missing):
                    link = p.get("link", "")
                    if link:
                        seller = fetch_seller_from_detail(driver, link)
                        p["penjual"] = seller if seller else "Seller Individu"
                    else:
                        p["penjual"] = "Seller Individu"
                    time.sleep(2)

    except Exception as e:
        send_ws_message(job_id, "error", {"message": "[BLIBLI] Error: {}".format(str(e)[:80])})
    finally:
        try:
            driver.quit()
        except Exception:
            pass

    # Export
    if all_products:
        send_ws_message(job_id, "status", {"message": "[BLIBLI] Mengexport {} produk ke Excel...".format(len(all_products))})
        hasil_dir = str(Path(__file__).parent.parent / "hasil" / "blibli")
        filepath = export_with_analytics(all_products, keyword, filters, output_dir=hasil_dir)
        return {"total": len(all_products), "file": os.path.basename(filepath)}

    return None


def run_shopee_job(job_id, keyword, pages, mode, filters):
    """Run Shopee scraping job - same flow as test_network_intercept.py"""
    import time
    import subprocess
    import undetected_chromedriver as uc
    from urllib.parse import quote_plus
    import re

    PROFILE_DIR = str(Path(__file__).parent.parent / "scrapers" / "shopee_profile")

    # Step 0: Kill existing Chrome to avoid profile lock (same as test script)
    send_ws_message(job_id, "status", {"message": "[SHOPEE] Menutup Chrome yang aktif (menghindari profile lock)..."})
    try:
        subprocess.run(["taskkill", "/F", "/IM", "chrome.exe", "/T"],
                      capture_output=True, timeout=10)
        time.sleep(3)
    except Exception:
        pass

    # Check if profile has valid login
    os.makedirs(PROFILE_DIR, exist_ok=True)
    cookies_file = os.path.join(PROFILE_DIR, "Default", "Cookies")
    first_time = not os.path.exists(cookies_file)

    if first_time:
        send_ws_message(job_id, "status", {"message": "[SHOPEE] Belum ada session login. Memulai proses login..."})
    else:
        send_ws_message(job_id, "status", {"message": "[SHOPEE] Session ditemukan. Membuka Chrome..."})

    # Open Chrome with profile (same options as test_network_intercept.py)
    send_ws_message(job_id, "progress", {"message": "[SHOPEE] Membuka Chrome browser..."})
    options = uc.ChromeOptions()
    options.add_argument("--start-maximized")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--user-data-dir={}".format(PROFILE_DIR))
    options.set_capability("goog:loggingPrefs", {"performance": "ALL"})

    try:
        driver = uc.Chrome(options=options, use_subprocess=True, version_main=147)
    except Exception as e:
        send_ws_message(job_id, "error", {"message": "[SHOPEE] Gagal buka Chrome: {}".format(str(e)[:50])})
        return None

    driver.set_page_load_timeout(90)
    all_products = []

    try:
        # ============ LOGIN FLOW (only if first time) ============
        if first_time:
            # Open Shopee login page directly - user does Google + Shopee login in one go
            send_ws_message(job_id, "progress", {"message": "[SHOPEE] [Step 1/2] Membuka halaman login Shopee..."})
            try:
                driver.get("https://shopee.co.id/buyer/login")
            except Exception:
                pass
            time.sleep(5)

            send_ws_message(job_id, "need_action", {
                "message": "[SHOPEE] [Step 1/2] Login di browser yang terbuka:\n1. Pilih metode login (Google / SMS / QR Code / Password)\n2. Selesaikan proses login\n3. Jika ada CAPTCHA, selesaikan puzzle verifikasi\n4. Tunggu sampai masuk ke halaman utama Shopee\n\nKlik 'Konfirmasi' setelah berhasil masuk ke Shopee.",
                "action": "login_shopee",
            })
            if not wait_for_confirmation(job_id, 600):  # 10 minutes for full login
                send_ws_message(job_id, "error", {"message": "[SHOPEE] Timeout login (10 menit)."})
                driver.quit()
                return None

            send_ws_message(job_id, "progress", {"message": "[SHOPEE] [Step 1/2] Login berhasil! ✓"})
            time.sleep(2)

            # After confirmation, navigate to Shopee homepage to save cookies
            send_ws_message(job_id, "status", {"message": "[SHOPEE] Memverifikasi session & menyimpan cookies..."})
            try:
                driver.switch_to.window(driver.window_handles[0])
            except Exception:
                pass
            try:
                driver.get("https://shopee.co.id/")
            except Exception:
                pass
            time.sleep(5)

            cookies = driver.get_cookies()
            shopee_cookies = [c["name"] for c in cookies if "SPC" in c.get("name", "")]
            if shopee_cookies:
                send_ws_message(job_id, "progress", {"message": "[SHOPEE] ✓ Cookies tersimpan: {}".format(", ".join(shopee_cookies[:3]))})
            else:
                send_ws_message(job_id, "status", {"message": "[SHOPEE] ⚠️ Cookies belum terdeteksi, tetap lanjut..."})
            time.sleep(2)

            # Close & reopen browser (clean state, same as test script)
            send_ws_message(job_id, "status", {"message": "[SHOPEE] Restart browser dengan session tersimpan..."})
            try:
                driver.quit()
            except Exception:
                pass
            time.sleep(3)

            # Create fresh options (cannot reuse ChromeOptions object)
            options2 = uc.ChromeOptions()
            options2.add_argument("--start-maximized")
            options2.add_argument("--no-sandbox")
            options2.add_argument("--disable-dev-shm-usage")
            options2.add_argument("--user-data-dir={}".format(PROFILE_DIR))
            options2.set_capability("goog:loggingPrefs", {"performance": "ALL"})

            try:
                driver = uc.Chrome(options=options2, use_subprocess=True, version_main=147)
                driver.set_page_load_timeout(90)
            except Exception as e:
                send_ws_message(job_id, "error", {"message": "[SHOPEE] Gagal restart Chrome: {}".format(str(e)[:50])})
                return None

            send_ws_message(job_id, "progress", {"message": "[SHOPEE] ✓ Browser siap dengan session login."})
            time.sleep(2)

        # ============ SCRAPING (same flow as test_network_intercept.py) ============
        send_ws_message(job_id, "progress", {"message": "[SHOPEE] [Step 2/2] Memulai scraping..."})
        keyword_encoded = quote_plus(keyword)

        for page_num in range(pages):
            url = "https://shopee.co.id/search?keyword={}".format(keyword_encoded)
            if page_num > 0:
                url += "&page={}".format(page_num)
            if filters.get("min_price"):
                url += "&minPrice={}".format(filters["min_price"])
            if filters.get("max_price"):
                url += "&maxPrice={}".format(filters["max_price"])
            if filters.get("rating"):
                url += "&ratingFilter={}".format(filters["rating"])
            if filters.get("sort"):
                url += "&sortBy={}".format(filters["sort"])

            send_ws_message(job_id, "progress", {
                "message": "[SHOPEE] 📄 Page {}/{}: Memuat halaman pencarian...".format(page_num + 1, pages),
                "page": page_num + 1,
            })

            # Clear performance logs before navigation
            try:
                driver.get_log("performance")
            except Exception:
                pass

            # Navigate to search page
            try:
                driver.get(url)
            except Exception:
                send_ws_message(job_id, "status", {"message": "[SHOPEE] ⚠️ Navigate timeout, menunggu..."})
                time.sleep(5)

            time.sleep(10)

            # Check CAPTCHA only on first page
            if page_num == 0:
                page_src = driver.page_source
                current_url = driver.current_url
                if any(x in page_src.lower() or x in current_url.lower() for x in ["verify", "captcha", "challenge"]):
                    send_ws_message(job_id, "need_action", {
                        "message": "[SHOPEE] CAPTCHA terdeteksi! Selesaikan puzzle di browser, lalu klik 'Konfirmasi'.",
                        "action": "solve_captcha",
                    })
                    if not wait_for_confirmation(job_id, 300):
                        send_ws_message(job_id, "error", {"message": "[SHOPEE] Timeout CAPTCHA (5 menit)."})
                        break
                    send_ws_message(job_id, "progress", {"message": "[SHOPEE] ✓ CAPTCHA berhasil! Melanjutkan..."})
                    time.sleep(3)
                    # Reload page after CAPTCHA (same as test script)
                    try:
                        driver.get_log("performance")
                        driver.get(url)
                        time.sleep(10)
                    except Exception:
                        pass
                else:
                    send_ws_message(job_id, "progress", {"message": "[SHOPEE] ✓ Tidak ada CAPTCHA. Lanjut scraping..."})

            # Scroll to trigger API calls (same as test script)
            send_ws_message(job_id, "status", {"message": "[SHOPEE] 📄 Page {}/{}: Scrolling & intercept API...".format(page_num + 1, pages)})
            try:
                for i in range(6):
                    driver.execute_script("window.scrollTo(0, document.body.scrollHeight * {}/100);".format((i+1)*16))
                    time.sleep(2)
                time.sleep(5)
            except Exception:
                pass

            # Extract from network logs (CDP method - same as test script)
            send_ws_message(job_id, "status", {"message": "[SHOPEE] 📄 Page {}/{}: Extracting data dari network...".format(page_num + 1, pages)})
            page_products = _extract_shopee_from_network(driver)

            # Fallback: JS DOM extraction if network empty
            if not page_products:
                send_ws_message(job_id, "status", {"message": "[SHOPEE] ⚠️ Network kosong, mencoba JS DOM extraction..."})
                page_products = _extract_shopee_via_js(driver)

            all_products.extend(page_products)
            send_ws_message(job_id, "progress", {
                "message": "[SHOPEE] ✅ Page {}/{}: {} produk ditemukan (Total: {})".format(page_num + 1, pages, len(page_products), len(all_products)),
                "products": len(all_products),
            })

            if page_num < pages - 1:
                send_ws_message(job_id, "status", {"message": "[SHOPEE] Menunggu sebelum page berikutnya..."})
                time.sleep(5)

    except Exception as e:
        send_ws_message(job_id, "error", {"message": "[SHOPEE] Error: {}".format(str(e)[:80])})
    finally:
        try:
            driver.quit()
        except Exception:
            pass

    # Export with analytics
    if all_products:
        send_ws_message(job_id, "status", {"message": "[SHOPEE] Mengexport {} produk ke Excel...".format(len(all_products))})
        from exporters.excel_analytics import export_with_analytics
        hasil_dir = str(Path(__file__).parent.parent / "hasil" / "shopee")
        filepath = export_with_analytics(all_products, keyword, filters, output_dir=hasil_dir)
        return {"total": len(all_products), "file": os.path.basename(filepath)}

    return None


def _is_blibli_cf_page(sb):
    """Check if current page is Cloudflare challenge on Blibli."""
    try:
        page_src = sb.get_page_source()
        current_url = sb.get_current_url()
        cf_indicators = ["challenge", "cf-browser-verification", "turnstile",
                        "checking your browser", "just a moment", "ray id",
                        "cloudflare"]
        return any(x in page_src.lower() or x in current_url.lower() for x in cf_indicators)
    except Exception:
        return False


def _parse_blibli_product(box):
    """Parse a single Blibli product card from BeautifulSoup element."""
    import re
    product = {}

    # Name
    name_el = box.find("span", class_=re.compile(r"product.*title|els-product__title"))
    if not name_el:
        name_el = box.find("p", class_=re.compile(r"product.*name|title"))
    if name_el:
        product["nama_produk"] = name_el.get_text(strip=True)[:150]
    else:
        title = box.get("title") or box.get("aria-label") or ""
        if title:
            product["nama_produk"] = title[:150]
    if not product.get("nama_produk"):
        return None

    # Price
    price_el = box.find("div", class_=re.compile(r"fixed.*price|els-product__fixed-price"))
    if not price_el:
        price_el = box.find("span", class_=re.compile(r"price|harga"))
    if price_el:
        price_text = price_el.get_text(strip=True)
        product["harga"] = price_text
        digits = re.sub(r"[^\d]", "", price_text)
        if digits:
            product["harga_angka"] = int(digits)

    # Original price
    orig_el = box.find("span", class_=re.compile(r"slash.*price|original.*price|strikethrough"))
    if orig_el:
        product["harga_sebelum_diskon"] = orig_el.get_text(strip=True)

    # Rating
    rating_el = box.find("span", class_=re.compile(r"rating"))
    if rating_el:
        try:
            product["rating"] = float(rating_el.get_text(strip=True))
        except (ValueError, TypeError):
            pass

    # Sold
    sold_el = box.find("span", class_=re.compile(r"sold|terjual"))
    if sold_el:
        sold_text = sold_el.get_text(strip=True)
        sold_digits = re.sub(r"[^\d]", "", sold_text)
        if sold_digits:
            product["terjual"] = int(sold_digits)

    # Seller
    seller_el = box.find("span", class_=re.compile(r"merchant|seller|store"))
    if seller_el:
        product["penjual"] = seller_el.get_text(strip=True)

    # Location
    loc_el = box.find("span", class_=re.compile(r"location|kota"))
    if loc_el:
        product["kota"] = loc_el.get_text(strip=True)

    # Link
    href = box.get("href", "")
    if href:
        if not href.startswith("http"):
            href = "https://www.blibli.com" + href
        product["link"] = href

    # Image
    img_el = box.find("img")
    if img_el:
        product["gambar"] = img_el.get("src") or img_el.get("data-src") or ""

    # Badge
    badge_el = box.find("span", class_=re.compile(r"badge"))
    if badge_el:
        product["badge"] = badge_el.get_text(strip=True)

    return product


def _extract_shopee_from_network(driver):
    """Extract Shopee products from intercepted network API responses (CDP method)."""
    import json as json_mod
    products = []

    try:
        logs = driver.get_log("performance")
    except Exception:
        return []

    for entry in logs:
        try:
            log_data = json_mod.loads(entry["message"])
            message = log_data.get("message", {})

            if message.get("method") == "Network.responseReceived":
                response = message.get("params", {}).get("response", {})
                url = response.get("url", "")

                if "search_items" in url or "search/search_items" in url:
                    request_id = message.get("params", {}).get("requestId")
                    if request_id:
                        try:
                            body = driver.execute_cdp_cmd("Network.getResponseBody", {"requestId": request_id})
                            if body and body.get("body"):
                                data = json_mod.loads(body["body"])
                                items = _parse_shopee_api_items(data)
                                products.extend(items)
                        except Exception:
                            pass
        except Exception:
            continue

    return products


def _parse_shopee_api_items(data):
    """Parse product items from Shopee API JSON response."""
    products = []
    items = []

    if isinstance(data, dict):
        items = data.get("items", [])
        if not items and isinstance(data.get("data"), dict):
            items = data["data"].get("items", [])

    for item in items:
        product = {}
        item_data = item.get("item_basic", item)

        name = item_data.get("name", "") or item_data.get("item_name", "")
        if not name:
            continue
        product["nama_produk"] = name[:150]

        # Price (Shopee stores in units of 100000)
        price = item_data.get("price", 0)
        price_min = item_data.get("price_min", 0)
        actual_price = price // 100000 if price > 100000 else price
        if actual_price == 0 and price_min > 0:
            actual_price = price_min // 100000 if price_min > 100000 else price_min

        if actual_price > 0:
            product["harga_angka"] = actual_price
            product["harga"] = "Rp{:,.0f}".format(actual_price).replace(",", ".")

        # Original price
        price_before = item_data.get("price_before_discount", 0)
        if price_before > 0:
            actual_before = price_before // 100000 if price_before > 100000 else price_before
            if actual_before > actual_price:
                product["harga_sebelum_diskon"] = "Rp{:,.0f}".format(actual_before).replace(",", ".")

        # Sold
        sold = item_data.get("sold", 0) or item_data.get("historical_sold", 0)
        if sold > 0:
            product["terjual"] = sold

        # Rating
        rating = item_data.get("item_rating", {})
        if isinstance(rating, dict):
            avg = rating.get("rating_star", 0)
            if avg > 0:
                product["rating"] = round(avg, 1)

        # Location
        location = item_data.get("shop_location", "") or item_data.get("item_location", "")
        if location:
            product["kota"] = location

        # Seller
        shop_name = item_data.get("shop_name", "")
        if not shop_name:
            if item_data.get("is_official_shop"):
                shop_name = "Official Shop"
            elif item_data.get("shopee_verified"):
                shop_name = "Shopee Mall"
        if shop_name:
            product["penjual"] = shop_name

        # Image
        image = item_data.get("image", "")
        if image and not image.startswith("http"):
            image = "https://cf.shopee.co.id/file/" + image
        product["gambar"] = image

        # Rekomendasi fields (affiliator)
        item_id = item_data.get("itemid", "") or item_data.get("item_id", "")
        if item_id:
            product["item_id"] = str(item_id)

        # Stock
        stock = item_data.get("stock", 0)
        if stock is None:
            stock = 0
        if stock > 0:
            product["stock"] = int(stock)

        # Liked count (n_like)
        liked = item_data.get("liked_count", 0) or item_data.get("n_like", 0)
        if liked and liked > 0:
            product["liked_count"] = int(liked)

        # Comment count
        cmt = item_data.get("cmt_count", 0) or item_data.get("comment_count", 0)
        if cmt and cmt > 0:
            product["comment_count"] = int(cmt)

        # Free shipping
        if item_data.get("is_free_shipping", False):
            product["free_shipping"] = "Ya"

        # Seller type
        if item_data.get("is_official_shop", False):
            product["seller_type"] = "Official"
        elif item_data.get("shopee_verified", False):
            product["seller_type"] = "Verified"
        elif item_data.get("is_shopee_verified", False):
            product["seller_type"] = "Verified"
        else:
            # Check shop_location / is_official for affiliate detection
            if item_data.get("is_preferred_plus_seller", False):
                product["seller_type"] = "Preferred"
            elif item_data.get("is_mall", 0) or item_data.get("is_official", 0):
                product["seller_type"] = "Mall"

        # Flash sale
        if item_data.get("flash_sale", False) or item_data.get("is_flash_sale", False):
            product["flash_sale"] = "Ya"
        elif item_data.get("badge_interaction", {}).get("flash_sale", False):
            product["flash_sale"] = "Ya"

        # Link
        shop_id = item_data.get("shopid", "") or item_data.get("shop_id", "")
        if item_id and shop_id:
            slug = name.replace(" ", "-").replace("/", "-")[:80]
            product["link"] = "https://shopee.co.id/{}-i.{}.{}".format(slug, shop_id, item_id)

        products.append(product)

    return products


def _extract_shopee_via_js(driver):
    """Fallback: Extract products via JavaScript DOM (same as test_network_intercept.py)."""
    try:
        products = driver.execute_script("""
            var products = [];
            var links = document.querySelectorAll('a[href*="-i."]');
            var seen = new Set();
            for (var i = 0; i < links.length; i++) {
                var link = links[i];
                var href = link.getAttribute('href') || '';
                if (seen.has(href) || !href.match(/-i\\.\\d+\\.\\d+/)) continue;
                seen.add(href);
                var card = link;
                for (var j = 0; j < 8; j++) {
                    if (card.parentElement && card.parentElement.innerText &&
                        card.parentElement.innerText.length > 30 &&
                        card.parentElement.innerText.length < 3000) {
                        card = card.parentElement; break;
                    }
                    if (card.parentElement) card = card.parentElement;
                }
                var text = card.innerText || '';
                var lines = text.split('\\n').map(l => l.trim()).filter(l => l);
                var product = {};
                for (var k = 0; k < lines.length; k++) {
                    if (lines[k].length > 10 && !/^[\\d₫Rp$]/.test(lines[k]) &&
                        !lines[k].toLowerCase().includes('terjual') && !/%/.test(lines[k])) {
                        product.nama_produk = lines[k].substring(0, 150); break;
                    }
                }
                if (!product.nama_produk) continue;
                var allElements = card.querySelectorAll('*');
                for (var k = 0; k < allElements.length; k++) {
                    var el = allElements[k];
                    var ariaLabel = el.getAttribute('aria-label') || '';
                    if (ariaLabel && ariaLabel.match(/\\d/) && ariaLabel.length < 30) {
                        var priceMatch = ariaLabel.replace(/[^\\d]/g, '');
                        if (priceMatch && parseInt(priceMatch) > 1000) {
                            product.harga_angka = parseInt(priceMatch);
                            product.harga = 'Rp' + parseInt(priceMatch).toLocaleString('id-ID');
                            break;
                        }
                    }
                }
                for (var k = 0; k < lines.length; k++) {
                    if (lines[k].toLowerCase().includes('terjual')) {
                        var soldText = lines[k].toLowerCase().replace('terjual', '').replace('+', '').trim();
                        var mult = soldText.includes('rb') ? 1000 : 1;
                        soldText = soldText.replace('rb', '').replace(',', '.').trim();
                        var soldNum = parseFloat(soldText);
                        if (!isNaN(soldNum)) product.terjual = Math.round(soldNum * mult);
                        break;
                    }
                }
                var locKw = ['Jakarta', 'Bandung', 'Surabaya', 'Bekasi', 'Tangerang', 'Bogor'];
                for (var k = lines.length - 1; k >= 0; k--) {
                    if (lines[k].length < 40 && locKw.some(l => lines[k].includes(l))) {
                        product.kota = lines[k]; break;
                    }
                }
                product.link = href.startsWith('http') ? href : 'https://shopee.co.id' + href;
                var img = link.querySelector('img');
                if (img) product.gambar = img.src || '';
                products.push(product);
            }
            return products;
        """)
        return products or []
    except Exception:
        return []


def _extract_shopee_product_old(link_el):
    """OLD METHOD - Extract Shopee product from link element (fallback)"""
    import re
    card = link_el
    for _ in range(6):
        if card.parent:
            text = card.parent.get_text(separator="|", strip=True)
            if "Rp" in text and len(text) > 30 and len(text) < 2000:
                card = card.parent
                break
            card = card.parent

    all_text = card.get_text(separator="|", strip=True)
    segments = [s.strip() for s in all_text.split("|") if s.strip()]
    product = {}

    for seg in segments:
        if (len(seg) > 10 and not seg.startswith("Rp") and
            "terjual" not in seg.lower() and "%" not in seg and
            not re.match(r"^\d+[.,]?\d*$", seg) and len(seg) < 200):
            product["nama_produk"] = seg[:150]
            break
    if not product.get("nama_produk"):
        return None

    for seg in segments:
        if seg.startswith("Rp"):
            cleaned = re.sub(r"[^\d]", "", seg)
            try:
                val = int(cleaned)
                if val > 1000:
                    product["harga_angka"] = val
                    product["harga"] = "Rp{:,.0f}".format(val).replace(",", ".")
                    break
            except ValueError:
                pass

    for seg in segments:
        if "terjual" in seg.lower():
            sold_text = seg.lower().replace("terjual", "").replace("+", "").strip()
            multiplier = 1000 if "rb" in sold_text else 1
            sold_text = sold_text.replace("rb", "").replace(",", ".").strip()
            try:
                product["terjual"] = int(float(sold_text) * multiplier)
            except (ValueError, TypeError):
                pass
            break

    for seg in segments:
        if re.match(r"^\d\.\d$", seg):
            try:
                val = float(seg)
                if 0 < val <= 5:
                    product["rating"] = val
                    break
            except ValueError:
                pass

    loc_kw = ["Jakarta", "Bandung", "Surabaya", "Bekasi", "Tangerang", "Semarang", "Depok", "Bogor"]
    for seg in segments:
        if any(loc in seg for loc in loc_kw) and len(seg) < 40:
            product["kota"] = seg
            break

    href = link_el.get("href", "")
    if not href.startswith("http"):
        href = "https://shopee.co.id" + href
    product["link"] = href

    return product


def _export_shopee_excel(products, keyword):
    """Export Shopee results to Excel"""
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    import re

    df = pd.DataFrame(products)
    col_order = ["nama_produk", "harga", "harga_angka", "terjual", "rating", "kota", "link", "page", "keyword"]
    existing = [c for c in col_order if c in df.columns]
    df = df[existing]

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_kw = re.sub(r"[^\w\s-]", "", keyword).replace(" ", "_")
    filename = "shopee_{}_{}.xlsx".format(safe_kw, timestamp)
    hasil_dir = Path(__file__).parent.parent / "hasil" / "shopee"
    hasil_dir.mkdir(parents=True, exist_ok=True)
    filepath = str(hasil_dir / filename)

    df.to_excel(filepath, index=False, sheet_name="Products", engine="openpyxl")

    wb = load_workbook(filepath)
    ws = wb.active
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill(start_color="EE4D2D", end_color="EE4D2D", fill_type="solid")
        cell.border = border
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 15
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).border = border
    for col in range(1, ws.max_column + 1):
        mx = max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, min(ws.max_row+1, 30)))
        ws.column_dimensions[get_column_letter(col)].width = min(mx + 2, 50)
    wb.save(filepath)
    return filepath


# ============================================================
# RUN SERVER
# ============================================================

if __name__ == "__main__":
    print("=" * 50)
    print("  Scraping Tools Web UI")
    print("  http://localhost:8080")
    print("=" * 50)
    uvicorn.run(app, host="127.0.0.1", port=9000)
