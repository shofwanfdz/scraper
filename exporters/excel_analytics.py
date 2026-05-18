"""
Excel Analytics Exporter - Professional Layout
Layout: Data tables on LEFT (col A-C), Charts on RIGHT (col E+) aligned with data.
Each section has proper spacing between them.
"""
import re
import os
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, ScatterChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart import Series
from openpyxl.utils import get_column_letter

# Layout constants
CHART_COL = "E"  # Charts start at column E (right of data)
CHART_WIDTH = 16
CHART_HEIGHT = 10
CHART_ROWS = 16  # How many rows a chart occupies visually
SECTION_GAP = 3  # Rows gap between sections

# Styles
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
TITLE_FONT = Font(bold=True, size=12, color="1F4E79")
SUBTITLE_FONT = Font(bold=True, size=10, color="2E75B6")
BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)


def _safe_int(val):
    """Safely convert value to int. Returns 0 for empty/invalid values."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return 0
    # Handle numeric types directly (avoid str(46.0) -> "460" bug)
    if isinstance(val, (int, float)):
        return int(val)
    # For strings: parse "Rp1.234.567", "100+", "1rb+", etc.
    s = str(val).strip()
    if not s:
        return 0
    # Handle "rb" (ribu/thousand) and "jt" (juta/million) suffixes
    lower = s.lower().replace("terjual", "").strip()
    multiplier = 1
    if "rb" in lower:
        multiplier = 1000
        lower = lower.replace("rb", "").strip()
    elif "jt" in lower:
        multiplier = 1000000
        lower = lower.replace("jt", "").strip()
    # Handle comma as decimal separator (Indonesian format: "4,2" = 4.2)
    lower = lower.replace(".", "").replace(",", ".").replace("+", "").strip()
    # Try to parse as float first (handles "4.2" correctly)
    try:
        return int(float(lower) * multiplier)
    except (ValueError, TypeError):
        pass
    # Final fallback: extract digits only
    digits = re.sub(r"[^\d]", "", str(val))
    return int(digits) if digits else 0


def _title(ws, row, text):
    c = ws.cell(row=row, column=1, value=text)
    c.font = TITLE_FONT
    return row + 1


def _subtitle(ws, row, text):
    """Write italic subtitle text, returns incremented row."""
    c = ws.cell(row=row, column=1, value=text)
    c.font = SUBTITLE_FONT
    return row + 1


def _header(ws, row, headers):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = BORDER
        c.alignment = Alignment(horizontal="center")
    return row + 1


def _row(ws, row, values):
    for col, v in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=v)
        c.border = BORDER
    return row + 1


def _auto_width(ws, max_col=4):
    for col in range(1, max_col + 1):
        mx = 0
        letter = get_column_letter(col)
        for r in range(1, min(ws.max_row + 1, 80)):
            val = ws.cell(row=r, column=col).value
            if val:
                mx = max(mx, len(str(val)))
        ws.column_dimensions[letter].width = min(mx + 2, 40)


def _bar_chart(ws, title, data_col, cat_col, start_row, end_row, anchor_row, horizontal=False):
    """Place bar chart at column E, aligned with anchor_row"""
    chart = BarChart()
    chart.type = "bar" if horizontal else "col"
    chart.title = title
    chart.style = 10
    chart.width = CHART_WIDTH
    chart.height = CHART_HEIGHT
    # Show legend with category names
    chart.legend.position = "b"  # bottom
    # Show data labels on bars
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.dataLabels.showCatName = False
    data = Reference(ws, min_col=data_col, min_row=start_row, max_row=end_row)
    cats = Reference(ws, min_col=cat_col, min_row=start_row + 1, max_row=end_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "{}{}".format(CHART_COL, anchor_row))


def _pie_chart(ws, title, data_col, cat_col, start_row, end_row, anchor_row):
    """Place pie chart at column E, aligned with anchor_row"""
    pie = PieChart()
    pie.title = title
    pie.style = 10
    pie.width = CHART_WIDTH
    pie.height = CHART_HEIGHT
    # Show legend with category names + colors
    pie.legend.position = "b"  # bottom
    # Show labels: category name + percentage + value
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showVal = True
    pie.dataLabels.showCatName = True
    data = Reference(ws, min_col=data_col, min_row=start_row, max_row=end_row)
    cats = Reference(ws, min_col=cat_col, min_row=start_row + 1, max_row=end_row)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(cats)
    ws.add_chart(pie, "{}{}".format(CHART_COL, anchor_row))


def _scatter_chart(ws, title, x_col, y_col, start_row, end_row, anchor_row):
    """Place scatter chart at column E"""
    scatter = ScatterChart()
    scatter.title = title
    scatter.x_axis.title = "Harga (Rp)"
    scatter.y_axis.title = "Terjual"
    scatter.style = 10
    scatter.width = CHART_WIDTH
    scatter.height = CHART_HEIGHT
    scatter.legend.position = "b"
    x_ref = Reference(ws, min_col=x_col, min_row=start_row, max_row=end_row)
    y_ref = Reference(ws, min_col=y_col, min_row=start_row, max_row=end_row)
    s = Series(y_ref, x_ref, title="Produk")
    scatter.series.append(s)
    ws.add_chart(scatter, "{}{}".format(CHART_COL, anchor_row))


# ============================================================
# MAIN EXPORT
# ============================================================

def export_with_analytics(products, keyword, filters=None, output_dir="hasil/blibli"):
    df = pd.DataFrame(products)

    # Normalize terjual: could be int, "100+", "1rb+", or "" from API
    if "terjual" in df.columns:
        df["terjual"] = df["terjual"].apply(
            lambda x: _safe_int(x) if pd.notna(x) else None)
        df["terjual"] = pd.to_numeric(df["terjual"], errors="coerce")

    # Ensure numeric columns
    if "harga_angka" not in df.columns and "harga" in df.columns:
        df["harga_angka"] = df["harga"].apply(
            lambda x: _safe_int(x) if pd.notna(x) and str(x).strip() else None)
    df["harga_angka"] = pd.to_numeric(df["harga_angka"], errors="coerce")

    if "harga_sebelum_diskon" in df.columns:
        df["harga_asli_angka"] = df["harga_sebelum_diskon"].apply(
            lambda x: _safe_int(x) if pd.notna(x) and str(x).strip() else None)
    else:
        df["harga_asli_angka"] = None
    df["harga_asli_angka"] = pd.to_numeric(df["harga_asli_angka"], errors="coerce")

    df["diskon_persen"] = pd.Series(dtype="float64")
    mask = df["harga_angka"].notna() & df["harga_asli_angka"].notna() & (df["harga_asli_angka"] > 0)
    df.loc[mask, "diskon_persen"] = round(
        (1 - df.loc[mask, "harga_angka"] / df.loc[mask, "harga_asli_angka"]) * 100, 1)
    df["diskon_persen"] = pd.to_numeric(df["diskon_persen"], errors="coerce")

    # Filename - detect marketplace from output_dir
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_kw = re.sub(r"[^\w\s-]", "", keyword).replace(" ", "_")
    marketplace = "blibli"
    for mp in ["shopee", "lazada", "tokopedia", "tiktokshop"]:
        if mp in output_dir.lower():
            marketplace = mp
            break
    filename = "{}_{}_analytics_{}.xlsx".format(marketplace, safe_kw, timestamp)
    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, filename)

    # Normalize: API mode uses "nama", HTML/stealth mode uses "nama_produk"
    if "nama" in df.columns and "nama_produk" not in df.columns:
        df.rename(columns={"nama": "nama_produk"}, inplace=True)

    # Extract brand from product name
    df["brand"] = df["nama_produk"].apply(_extract_brand)

    # Sheet 1: Products — rename columns for display
    COL_RENAME = {
        "item_id": "SKU",
        "liked_count": "Total Favorit",
        "comment_count": "Jumlah Ulasan",
    }

    col_order = [
        # --- Field Wajib ---
        "nama_produk", "brand", "harga", "harga_angka", "harga_sebelum_diskon", "diskon_persen",
        "penjual", "kota", "terjual", "rating",
        # --- Field Rekomendasi (Affiliator) ---
        "item_id", "stock", "liked_count", "comment_count", "free_shipping", "seller_type",
        "flash_sale",
        # --- Field Lainnya ---
        "badge", "cicilan", "link", "gambar", "page", "keyword", "scrape_time"
    ]
    existing = [c for c in col_order if c in df.columns]
    df_export = df[existing].rename(columns=COL_RENAME)
    df_export.to_excel(filepath, index=False, sheet_name="Products", engine="openpyxl")

    wb = load_workbook(filepath)
    ws = wb["Products"]
    _header(ws, 1, [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)])

    # Find 'link' column index
    link_col = None
    for c in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=c).value == "link":
            link_col = c
            break

    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 15
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).border = BORDER
        # Convert link column to clickable hyperlink
        if link_col:
            cell = ws.cell(row=r, column=link_col)
            url = cell.value
            if url and isinstance(url, str) and url.startswith("http"):
                cell.hyperlink = url
                cell.value = url
                cell.font = Font(color="0563C1", underline="single", size=10)
    _auto_width(ws, ws.max_column)

    # Build sheets
    _build_harga(wb, df)
    _build_seller(wb, df)
    _build_lokasi(wb, df)
    _build_penjualan(wb, df)
    _build_diskon(wb, df)
    _build_brand(wb, df)
    _build_best_value(wb, df)
    _build_rekomendasi(wb, df)
    _build_rekomendasi_v2(wb, df)
    _build_dashboard(wb, df, keyword, filters, marketplace)

    # Reorder: Dashboard after Products
    wb.move_sheet("Ecommerce Scraping", offset=-(len(wb.sheetnames) - 2))

    wb.save(filepath)
    return filepath


# ============================================================
# ANALISIS HARGA
# ============================================================

def _build_harga(wb, df):
    ws = wb.create_sheet("Analisis Harga")
    harga = df["harga_angka"].dropna()
    if len(harga) == 0:
        return

    row = 1

    # Section 1: Distribusi Harga (data left, chart right)
    row = _title(ws, row, "DISTRIBUSI HARGA")
    min_p, max_p = int(harga.min()), int(harga.max())
    step = max((max_p - min_p) // 6, 500000)
    ranges = []
    for i in range(min_p, max_p + step, step):
        upper = min(i + step, max_p + 1)
        count = int(((harga >= i) & (harga < upper)).sum())
        label = "Rp{:,.0f} - Rp{:,.0f}".format(i, upper - 1)
        ranges.append((label, count))
    ranges = [r for r in ranges[:8] if r[1] > 0]

    data_start = row
    row = _header(ws, row, ["Range Harga", "Jumlah"])
    for label, count in ranges:
        row = _row(ws, row, [label, count])

    # Chart on the RIGHT at same starting row
    _bar_chart(ws, "Distribusi Harga Produk", 2, 1, data_start, data_start + len(ranges), data_start)

    # Gap
    row = max(row, data_start + CHART_ROWS + 3) + SECTION_GAP

    # Section 2: Top 10 Termurah
    row = _title(ws, row, "TOP 10 PRODUK TERMURAH")
    row = _header(ws, row, ["Nama Produk", "Harga", "Penjual"])
    for _, r in df.nsmallest(10, "harga_angka").iterrows():
        row = _row(ws, row, [str(r.get("nama_produk", ""))[:55], str(r.get("harga", "")), str(r.get("penjual", ""))[:30]])

    row += SECTION_GAP

    # Section 3: Top 10 Termahal
    row = _title(ws, row, "TOP 10 PRODUK TERMAHAL")
    row = _header(ws, row, ["Nama Produk", "Harga", "Penjual"])
    for _, r in df.nlargest(10, "harga_angka").iterrows():
        row = _row(ws, row, [str(r.get("nama_produk", ""))[:55], str(r.get("harga", "")), str(r.get("penjual", ""))[:30]])

    row += SECTION_GAP

    # Section 4: Harga rata-rata per lokasi (detail analysis)
    row = _title(ws, row, "HARGA RATA-RATA PER LOKASI")
    avg_loc = df.groupby("kota")["harga_angka"].agg(["mean", "min", "max", "count"]).sort_values("mean", ascending=False)
    data_start2 = row
    row = _header(ws, row, ["Kota", "Rata-rata", "Termurah", "Termahal", "Jml Produk"])
    for kota, r in avg_loc.iterrows():
        row = _row(ws, row, [
            str(kota), "Rp{:,.0f}".format(r["mean"]),
            "Rp{:,.0f}".format(r["min"]), "Rp{:,.0f}".format(r["max"]), int(r["count"])
        ])

    _bar_chart(ws, "Harga Rata-rata per Lokasi", 2, 1, data_start2, data_start2 + len(avg_loc), data_start2, horizontal=True)

    _auto_width(ws, 5)


# ============================================================
# ANALISIS SELLER
# ============================================================

def _build_seller(wb, df):
    ws = wb.create_sheet("Analisis Seller")
    if "penjual" not in df.columns:
        return

    row = 1

    # Section 1: Top 10 Seller (data left, chart right)
    row = _title(ws, row, "TOP 10 SELLER (JUMLAH PRODUK)")
    top = df["penjual"].value_counts().head(10)
    data_start = row
    row = _header(ws, row, ["Seller", "Jumlah Produk"])
    for seller, count in top.items():
        row = _row(ws, row, [str(seller)[:35], int(count)])

    _bar_chart(ws, "Top 10 Seller", 2, 1, data_start, data_start + len(top), data_start, horizontal=True)

    row = max(row, data_start + CHART_ROWS + 3) + SECTION_GAP

    # Section 2: Tipe Seller (data left, pie right)
    row = _title(ws, row, "DISTRIBUSI TIPE SELLER")

    if "seller_type" in df.columns and df["seller_type"].notna().any():
        # Gunakan field seller_type yang baru (dari API)
        seller_type_counts = df["seller_type"].value_counts()
        data_start = row
        row = _header(ws, row, ["Tipe Seller", "Jumlah"])
        for stype, count in seller_type_counts.items():
            row = _row(ws, row, [str(stype), int(count)])
        if len(seller_type_counts) > 1:
            _pie_chart(ws, "Distribusi Tipe Seller", 2, 1, data_start, data_start + len(seller_type_counts), data_start)
    else:
        # Fallback: deteksi dari nama seller
        flagship = int(df["penjual"].str.contains("Flagship|Official|Blibli|Shopee|Mall", case=False, na=False).sum())
        individu = int((df["penjual"] == "Seller Individu").sum())
        regular = len(df) - flagship - individu

        data_start = row
        row = _header(ws, row, ["Tipe Seller", "Jumlah"])
        row = _row(ws, row, ["Flagship/Official", flagship])
        row = _row(ws, row, ["Seller Regular", regular])
        row = _row(ws, row, ["Seller Individu", individu])

        _pie_chart(ws, "Distribusi Tipe Seller", 2, 1, data_start, data_start + 3, data_start)

    row = max(row, data_start + CHART_ROWS + 3) + SECTION_GAP

    # Section 3: Avg Rating per Seller
    if "rating" in df.columns:
        row = _title(ws, row, "RATA-RATA RATING PER SELLER")
        avg_rat = df.groupby("penjual")["rating"].mean().sort_values(ascending=False).head(10)
        data_start = row
        row = _header(ws, row, ["Seller", "Avg Rating"])
        for seller, rating in avg_rat.items():
            row = _row(ws, row, [str(seller)[:35], round(float(rating), 2)])

        _bar_chart(ws, "Rating per Seller", 2, 1, data_start, data_start + len(avg_rat), data_start, horizontal=True)

        row = max(row, data_start + CHART_ROWS + 3) + SECTION_GAP

    # Section 4: Seller dengan penjualan tertinggi
    if "terjual" in df.columns and df["terjual"].notna().any():
        row = _title(ws, row, "SELLER DENGAN TOTAL PENJUALAN TERTINGGI")
        seller_sales = df.groupby("penjual")["terjual"].sum().sort_values(ascending=False).head(10)
        data_start = row
        row = _header(ws, row, ["Seller", "Total Terjual"])
        for seller, total in seller_sales.items():
            row = _row(ws, row, [str(seller)[:35], int(total)])

        _bar_chart(ws, "Total Penjualan per Seller", 2, 1, data_start, data_start + len(seller_sales), data_start, horizontal=True)

    _auto_width(ws, 3)


# ============================================================
# ANALISIS LOKASI
# ============================================================

def _build_lokasi(wb, df):
    ws = wb.create_sheet("Analisis Lokasi")
    if "kota" not in df.columns:
        return

    row = 1

    # Section 1: Distribusi per Kota
    row = _title(ws, row, "DISTRIBUSI PRODUK PER KOTA")
    city_counts = df["kota"].value_counts().head(10)
    data_start = row
    row = _header(ws, row, ["Kota", "Jumlah Produk"])
    for city, count in city_counts.items():
        row = _row(ws, row, [str(city), int(count)])

    _bar_chart(ws, "Distribusi Produk per Kota", 2, 1, data_start, data_start + len(city_counts), data_start, horizontal=True)

    row = max(row, data_start + CHART_ROWS + 3) + SECTION_GAP

    # Section 2: Harga rata-rata per kota (detail)
    row = _title(ws, row, "DETAIL HARGA PER KOTA")
    avg_price = df.groupby("kota")["harga_angka"].agg(["mean", "min", "max", "count"]).sort_values("mean", ascending=False)
    data_start = row
    row = _header(ws, row, ["Kota", "Rata-rata", "Min", "Max", "Produk"])
    for kota, r in avg_price.iterrows():
        row = _row(ws, row, [
            str(kota), "Rp{:,.0f}".format(r["mean"]),
            "Rp{:,.0f}".format(r["min"]), "Rp{:,.0f}".format(r["max"]), int(r["count"])
        ])

    _bar_chart(ws, "Rata-rata Harga per Kota", 2, 1, data_start, data_start + len(avg_price), data_start)

    row = max(row, data_start + CHART_ROWS + 3) + SECTION_GAP

    # Section 3: Seller terbanyak per kota
    row = _title(ws, row, "SELLER TERBANYAK PER KOTA")
    if "penjual" in df.columns:
        city_seller = df.groupby("kota")["penjual"].nunique().sort_values(ascending=False).head(10)
        data_start = row
        row = _header(ws, row, ["Kota", "Jumlah Seller Unik"])
        for kota, count in city_seller.items():
            row = _row(ws, row, [str(kota), int(count)])

        _bar_chart(ws, "Jumlah Seller per Kota", 2, 1, data_start, data_start + len(city_seller), data_start, horizontal=True)

    _auto_width(ws, 5)


# ============================================================
# ANALISIS PENJUALAN
# ============================================================

def _build_penjualan(wb, df):
    ws = wb.create_sheet("Analisis Penjualan")
    row = 1

    # Section 1: Top 10 Terlaris
    row = _title(ws, row, "TOP 10 PRODUK TERLARIS")
    if "terjual" in df.columns and df["terjual"].notna().any():
        terlaris = df[df["terjual"].notna()].nlargest(10, "terjual")
        data_start = row
        row = _header(ws, row, ["Nama Produk", "Harga", "Terjual", "Rating"])
        for _, r in terlaris.iterrows():
            row = _row(ws, row, [
                str(r.get("nama_produk", ""))[:50], str(r.get("harga", "")),
                _safe_int(r["terjual"]),
                float(r["rating"]) if pd.notna(r.get("rating")) else "",
            ])

        _bar_chart(ws, "Top 10 Terlaris", 3, 1, data_start, data_start + len(terlaris), data_start, horizontal=True)
    else:
        _row(ws, row, ["Tidak ada data penjualan"])
        data_start = row

    row = max(row, data_start + CHART_ROWS + 3) + SECTION_GAP

    # Section 2: Distribusi Rating
    if "rating" in df.columns and df["rating"].notna().any():
        row = _title(ws, row, "DISTRIBUSI RATING")
        rating_counts = df["rating"].value_counts().sort_index()
        data_start = row
        row = _header(ws, row, ["Rating", "Jumlah"])
        for rating, count in rating_counts.items():
            row = _row(ws, row, [float(rating) if pd.notna(rating) else "N/A", int(count)])

        _pie_chart(ws, "Distribusi Rating", 2, 1, data_start, data_start + len(rating_counts), data_start)

        row = max(row, data_start + CHART_ROWS + 3) + SECTION_GAP

    # Section 3: Korelasi Harga vs Terjual
    if "terjual" in df.columns and "harga_angka" in df.columns:
        scatter_df = df[df["terjual"].notna() & df["harga_angka"].notna()][["harga_angka", "terjual"]].head(50)
        if len(scatter_df) > 3:
            row = _title(ws, row, "KORELASI HARGA VS TERJUAL")
            data_start = row
            row = _header(ws, row, ["Harga (Rp)", "Terjual"])
            for _, r in scatter_df.iterrows():
                row = _row(ws, row, [_safe_int(r["harga_angka"]), _safe_int(r["terjual"])])

            _scatter_chart(ws, "Harga vs Terjual", 1, 2, data_start + 1, data_start + len(scatter_df), data_start)

    _auto_width(ws, 4)


# ============================================================
# ANALISIS DISKON
# ============================================================

def _build_diskon(wb, df):
    ws = wb.create_sheet("Analisis Diskon")
    row = 1

    # Computed in export_with_analytics, but handle if not present
    if "diskon_persen" not in df.columns:
        if "harga_sebelum_diskon" in df.columns and "harga_angka" in df.columns:
            df = df.copy()
            df["harga_asli_angka"] = df["harga_sebelum_diskon"].apply(
                lambda x: int(d) if pd.notna(x) and str(x).strip() and (d := re.sub(r"[^\d]", "", str(x))) else None)
            df["harga_asli_angka"] = pd.to_numeric(df["harga_asli_angka"], errors="coerce")
            mask = df["harga_angka"].notna() & df["harga_asli_angka"].notna() & (df["harga_asli_angka"] > 0)
            df["diskon_persen"] = pd.Series(dtype="float64")
            df.loc[mask, "diskon_persen"] = round(
                (1 - df.loc[mask, "harga_angka"] / df.loc[mask, "harga_asli_angka"]) * 100, 1)
            df["diskon_persen"] = pd.to_numeric(df["diskon_persen"], errors="coerce")
        else:
            ws.cell(row=1, column=1, value="Tidak ada data diskon").font = Font(italic=True, color="888888")
            _auto_width(ws, 4)
            return

    has_diskon = df["diskon_persen"].notna()
    df_diskon = df[has_diskon].copy()
    count_diskon = int(has_diskon.sum())
    count_no = len(df) - count_diskon

    # Section 1: Proporsi Diskon
    row = _title(ws, row, "PROPORSI PRODUK DISKON")
    data_start = row
    row = _header(ws, row, ["Status", "Jumlah"])
    row = _row(ws, row, ["Diskon", count_diskon])
    row = _row(ws, row, ["Tanpa Diskon", count_no])

    _pie_chart(ws, "Diskon vs Tanpa Diskon", 2, 1, data_start, data_start + 2, data_start)

    row = max(row, data_start + CHART_ROWS + 3) + SECTION_GAP

    # Section 2: Top 10 Diskon Terbesar
    row = _title(ws, row, "TOP 10 DISKON TERBESAR")
    if len(df_diskon) > 0:
        top_d = df_diskon.nlargest(10, "diskon_persen")
        row = _header(ws, row, ["Nama Produk", "Harga", "Harga Asli", "Diskon %"])
        for _, r in top_d.iterrows():
            row = _row(ws, row, [
                str(r.get("nama_produk", ""))[:50], str(r.get("harga", "")),
                str(r.get("harga_sebelum_diskon", "")),
                "{:.1f}%".format(r["diskon_persen"]) if pd.notna(r["diskon_persen"]) else "",
            ])
    else:
        row = _row(ws, row, ["Tidak ada produk diskon"])

    row += SECTION_GAP

    # Section 3: Avg Diskon per Seller
    if len(df_diskon) > 0:
        row = _title(ws, row, "RATA-RATA DISKON PER SELLER")
        avg_d = df_diskon.groupby("penjual")["diskon_persen"].mean().sort_values(ascending=False).head(10)
        data_start = row
        row = _header(ws, row, ["Seller", "Avg Diskon %"])
        for seller, avg in avg_d.items():
            row = _row(ws, row, [str(seller)[:35], round(float(avg), 1)])

        _bar_chart(ws, "Rata-rata Diskon per Seller", 2, 1, data_start, data_start + len(avg_d), data_start, horizontal=True)

    _auto_width(ws, 4)


# ============================================================
# BRAND EXTRACTION
# ============================================================

KNOWN_BRANDS = [
    "ASUS", "HP", "Lenovo", "Acer", "Dell", "MSI", "Apple", "MacBook",
    "Samsung", "Toshiba", "Axioo", "Polytron", "Infinix", "Realme",
    "Xiaomi", "OPPO", "Vivo", "iPhone", "Huawei", "Sony", "LG",
    "Adidas", "Nike", "Puma", "New Balance", "Asics", "Converse",
    "Logitech", "Razer", "SteelSeries", "Corsair", "HyperX",
    "Canon", "Nikon", "Fujifilm", "GoPro", "DJI",
    "Philips", "Panasonic", "Sharp", "Daikin", "Electrolux",
]


_DB_BRAND_CACHE = None  # None=uninitialized, []=DB unavailable, list=loaded


def _load_db_brands_once():
    """Load DB brands exactly once; on failure mark as unavailable."""
    global _DB_BRAND_CACHE
    if _DB_BRAND_CACHE is not None:
        return _DB_BRAND_CACHE
    try:
        from core.brand_manager import BrandManager
        mgr = BrandManager()
        _DB_BRAND_CACHE = list(mgr.get_all_brands_from_db() or [])
    except Exception:
        _DB_BRAND_CACHE = []
    return _DB_BRAND_CACHE


def _extract_brand(name):
    """Extract brand name from product name. Uses DB brands + fallback list."""
    if not name or not isinstance(name, str):
        return "Lainnya"

    # Try DB once then cache result (avoids N retry loops when DB is down)
    db_brands = _load_db_brands_once()
    if db_brands:
        all_brands = db_brands + [b for b in KNOWN_BRANDS if b not in db_brands]
    else:
        all_brands = list(KNOWN_BRANDS)

    name_upper = name.upper()
    for brand in all_brands:
        if brand.upper() in name_upper:
            if brand.upper() in ("MACBOOK", "IPHONE"):
                return "Apple"
            return brand.upper() if len(brand) <= 4 else brand
    return "Lainnya"


# ============================================================
# ANALISIS BRAND
# ============================================================

def _build_brand(wb, df):
    ws = wb.create_sheet("Analisis Brand")
    if "brand" not in df.columns:
        return

    row = 1

    # Section 1: Market Share per Brand (pie)
    row = _title(ws, row, "MARKET SHARE PER BRAND")
    brand_counts = df["brand"].value_counts().head(10)
    data_start = row
    row = _header(ws, row, ["Brand", "Jumlah Produk"])
    for brand, count in brand_counts.items():
        row = _row(ws, row, [str(brand), int(count)])

    _pie_chart(ws, "Market Share per Brand", 2, 1, data_start, data_start + len(brand_counts), data_start)

    row = max(row, data_start + CHART_ROWS + 3) + SECTION_GAP

    # Section 2: Harga Rata-rata per Brand (bar)
    row = _title(ws, row, "HARGA RATA-RATA PER BRAND")
    avg_brand = df.groupby("brand")["harga_angka"].mean().sort_values(ascending=False).head(10)
    data_start = row
    row = _header(ws, row, ["Brand", "Rata-rata Harga"])
    for brand, avg in avg_brand.items():
        row = _row(ws, row, [str(brand), int(avg)])

    _bar_chart(ws, "Harga Rata-rata per Brand", 2, 1, data_start, data_start + len(avg_brand), data_start, horizontal=True)

    row = max(row, data_start + CHART_ROWS + 3) + SECTION_GAP

    # Section 3: Rating per Brand (bar)
    if "rating" in df.columns:
        row = _title(ws, row, "RATING RATA-RATA PER BRAND")
        avg_rating_brand = df.groupby("brand")["rating"].mean().sort_values(ascending=False).head(10)
        data_start = row
        row = _header(ws, row, ["Brand", "Avg Rating"])
        for brand, rating in avg_rating_brand.items():
            row = _row(ws, row, [str(brand), round(float(rating), 2)])

        _bar_chart(ws, "Rating per Brand", 2, 1, data_start, data_start + len(avg_rating_brand), data_start, horizontal=True)

        row = max(row, data_start + CHART_ROWS + 3) + SECTION_GAP

    # Section 4: Price Segment (pie)
    row = _title(ws, row, "SEGMENTASI HARGA")
    harga = df["harga_angka"].dropna()
    segments = {
        "Budget (< 5 Juta)": int((harga < 5000000).sum()),
        "Mid-Range (5-10 Juta)": int(((harga >= 5000000) & (harga < 10000000)).sum()),
        "High-End (10-20 Juta)": int(((harga >= 10000000) & (harga < 20000000)).sum()),
        "Premium (> 20 Juta)": int((harga >= 20000000).sum()),
    }
    # Remove zero segments
    segments = {k: v for k, v in segments.items() if v > 0}

    data_start = row
    row = _header(ws, row, ["Segmen", "Jumlah Produk"])
    for seg, count in segments.items():
        row = _row(ws, row, [seg, count])

    _pie_chart(ws, "Segmentasi Harga Produk", 2, 1, data_start, data_start + len(segments), data_start)

    row = max(row, data_start + CHART_ROWS + 3) + SECTION_GAP

    # Section 5: Detail Brand - Harga Min/Max/Avg per Brand
    row = _title(ws, row, "DETAIL HARGA PER BRAND")
    brand_detail = df.groupby("brand")["harga_angka"].agg(["count", "mean", "min", "max"]).sort_values("count", ascending=False)
    data_start = row
    row = _header(ws, row, ["Brand", "Produk", "Rata-rata", "Termurah", "Termahal"])
    for brand, r in brand_detail.iterrows():
        row = _row(ws, row, [
            str(brand), int(r["count"]),
            "Rp{:,.0f}".format(r["mean"]),
            "Rp{:,.0f}".format(r["min"]),
            "Rp{:,.0f}".format(r["max"]),
        ])

    _bar_chart(ws, "Jumlah Produk per Brand", 2, 1, data_start, data_start + len(brand_detail), data_start, horizontal=True)

    row = max(row, data_start + CHART_ROWS + 3) + SECTION_GAP

    row += SECTION_GAP

    # Section 6: Seller Termurah per Brand
    row = _title(ws, row, "SELLER TERMURAH PER BRAND")
    row = _header(ws, row, ["Brand", "Produk Termurah", "Harga", "Seller"])
    for brand in brand_detail.index:
        brand_df = df[df["brand"] == brand]
        if len(brand_df) > 0 and brand_df["harga_angka"].notna().any():
            cheapest = brand_df.loc[brand_df["harga_angka"].idxmin()]
            row = _row(ws, row, [
                str(brand),
                str(cheapest.get("nama_produk", ""))[:40],
                str(cheapest.get("harga", "")),
                str(cheapest.get("penjual", ""))[:25],
            ])

    _auto_width(ws, 5)


# ============================================================
# BEST VALUE PRODUCTS
# ============================================================

def _build_best_value(wb, df):
    ws = wb.create_sheet("Best Value")
    row = 1

    # Calculate value score: (rating * terjual) / (harga / 1000000)
    # Higher = better value
    df_scored = df.copy()
    df_scored["terjual_safe"] = pd.to_numeric(df_scored.get("terjual"), errors="coerce").fillna(0)
    df_scored["rating_safe"] = pd.to_numeric(df_scored.get("rating"), errors="coerce").fillna(0)
    df_scored["harga_safe"] = df_scored["harga_angka"].fillna(1)

    # Value score formula
    df_scored["value_score"] = round(
        (df_scored["rating_safe"] * (df_scored["terjual_safe"] + 1)) / (df_scored["harga_safe"] / 1000000), 2
    )

    # Section 1: Top 15 Best Value Products
    row = _title(ws, row, "TOP 15 PRODUK BEST VALUE")
    row = _subtitle(ws, row, "Skor = (Rating x Terjual) / (Harga/1jt) - Semakin tinggi semakin worth it")

    top_value = df_scored[df_scored["value_score"] > 0].nlargest(15, "value_score")
    data_start = row
    row = _header(ws, row, ["Nama Produk", "Harga", "Rating", "Terjual", "Value Score"])
    for _, r in top_value.iterrows():
        row = _row(ws, row, [
            str(r.get("nama_produk", ""))[:50],
            str(r.get("harga", "")),
            float(r["rating_safe"]) if r["rating_safe"] > 0 else "",
            int(r["terjual_safe"]) if r["terjual_safe"] > 0 else "",
            float(r["value_score"]),
        ])

    # Bar chart for value scores
    if len(top_value) > 0:
        _bar_chart(ws, "Top 15 Best Value Score", 5, 1, data_start, data_start + len(top_value), data_start, horizontal=True)

    row = max(row, data_start + CHART_ROWS + 3) + SECTION_GAP

    # Section 2: Best Value per Brand
    row = _title(ws, row, "BEST VALUE PER BRAND")
    row = _subtitle(ws, row, "Produk dengan value score tertinggi di setiap brand")
    row = _header(ws, row, ["Brand", "Produk", "Harga", "Rating", "Score"])

    if "brand" in df_scored.columns:
        for brand in df_scored["brand"].unique():
            brand_df = df_scored[df_scored["brand"] == brand]
            if len(brand_df) > 0 and brand_df["value_score"].max() > 0:
                best = brand_df.loc[brand_df["value_score"].idxmax()]
                row = _row(ws, row, [
                    str(brand),
                    str(best.get("nama_produk", ""))[:40],
                    str(best.get("harga", "")),
                    float(best["rating_safe"]) if best["rating_safe"] > 0 else "",
                    float(best["value_score"]),
                ])

    row += SECTION_GAP

    # Compute diskon_persen if not present (needed for scatter chart below)
    df_bv = df.copy()
    if "diskon_persen" not in df_bv.columns:
        if "harga_sebelum_diskon" in df_bv.columns and "harga_angka" in df_bv.columns:
            df_bv["harga_asli_angka"] = df_bv["harga_sebelum_diskon"].apply(
                lambda x: _safe_int(x) if pd.notna(x) and str(x).strip() else None)
            df_bv["harga_asli_angka"] = pd.to_numeric(df_bv["harga_asli_angka"], errors="coerce")
            df_bv["diskon_persen"] = pd.Series(dtype="float64")
            mask = df_bv["harga_angka"].notna() & df_bv["harga_asli_angka"].notna() & (df_bv["harga_asli_angka"] > 0)
            df_bv.loc[mask, "diskon_persen"] = round(
                (1 - df_bv.loc[mask, "harga_angka"] / df_bv.loc[mask, "harga_asli_angka"]) * 100, 1)
            df_bv["diskon_persen"] = pd.to_numeric(df_bv["diskon_persen"], errors="coerce")
        else:
            df_bv["diskon_persen"] = pd.Series(dtype="float64")

    # Section 3: Diskon vs Rating scatter
    if df_bv["diskon_persen"].notna().any() and "rating" in df_bv.columns:
        row = _title(ws, row, "KORELASI DISKON VS RATING")
        row = _subtitle(ws, row, "Apakah produk diskon besar memiliki rating rendah?")

        scatter_df = df_bv[df_bv["diskon_persen"].notna() & df_bv["rating"].notna()][["diskon_persen", "rating"]].head(50)
        if len(scatter_df) > 3:
            data_start = row
            row = _header(ws, row, ["Diskon %", "Rating"])
            for _, r in scatter_df.iterrows():
                row = _row(ws, row, [float(r["diskon_persen"]), float(r["rating"])])

            # Scatter chart
            scatter = ScatterChart()
            scatter.title = "Diskon vs Rating"
            scatter.x_axis.title = "Diskon (%)"
            scatter.y_axis.title = "Rating"
            scatter.style = 10
            scatter.width = CHART_WIDTH
            scatter.height = CHART_HEIGHT
            scatter.legend.position = "b"
            x_ref = Reference(ws, min_col=1, min_row=data_start + 1, max_row=data_start + len(scatter_df))
            y_ref = Reference(ws, min_col=2, min_row=data_start + 1, max_row=data_start + len(scatter_df))
            s = Series(y_ref, x_ref, title="Produk")
            scatter.series.append(s)
            ws.add_chart(scatter, "{}{}".format(CHART_COL, data_start))

    _auto_width(ws, 5)


# ============================================================
# ANALISIS FIELD REKOMENDASI (AFFILIATOR)
# ============================================================

def _build_rekomendasi(wb, df):
    ws = wb.create_sheet("Analisis Rekomendasi")
    row = 1

    has_data = any(c in df.columns for c in ["liked_count", "comment_count", "stock", "seller_type", "free_shipping", "flash_sale"])
    if not has_data:
        ws.cell(row=1, column=1, value="Tidak ada data field rekomendasi").font = Font(italic=True, color="888888")
        _auto_width(ws, 3)
        return

    # Section 1: Distribusi Tipe Seller
    if "seller_type" in df.columns and df["seller_type"].notna().any():
        row = _title(ws, row, "DISTRIBUSI TIPE SELLER")
        seller_type_counts = df["seller_type"].value_counts()
        data_start = row
        row = _header(ws, row, ["Tipe Seller", "Jumlah"])
        for stype, count in seller_type_counts.items():
            row = _row(ws, row, [str(stype), int(count)])
        if len(seller_type_counts) > 1:
            _pie_chart(ws, "Distribusi Tipe Seller", 2, 1, data_start, data_start + len(seller_type_counts), data_start)
        row = max(row, data_start + CHART_ROWS + 3) + SECTION_GAP

    # Section 2: Free Shipping Analysis
    if "free_shipping" in df.columns:
        row = _title(ws, row, "FREE SHIPPING vs TANPA FREE SHIPPING")
        free = int((df["free_shipping"] == "Ya").sum())
        not_free = len(df) - free
        data_start = row
        row = _header(ws, row, ["Status", "Jumlah"])
        row = _row(ws, row, ["Free Shipping", free])
        row = _row(ws, row, ["Tanpa Free Shipping", not_free])
        _pie_chart(ws, "Free Shipping", 2, 1, data_start, data_start + 2, data_start)
        row = max(row, data_start + CHART_ROWS + 3) + SECTION_GAP

    # Section 3: Flash Sale Analysis
    if "flash_sale" in df.columns and df["flash_sale"].notna().any():
        row = _title(ws, row, "PRODUK FLASH SALE")
        flash = df[df["flash_sale"] == "Ya"]
        if len(flash) > 0:
            data_start = row
            row = _header(ws, row, ["Nama Produk", "Harga", "Terjual", "Seller"])
            for _, r in flash.iterrows():
                row = _row(ws, row, [
                    str(r.get("nama_produk", ""))[:50],
                    str(r.get("harga", "")),
                    int(r["terjual"]) if pd.notna(r.get("terjual")) else "",
                    str(r.get("penjual", ""))[:25],
                ])
        else:
            row = _row(ws, row, ["Tidak ada produk flash sale"])
        row += SECTION_GAP

    # Section 4: Top 10 by Liked Count
    if "liked_count" in df.columns and df["liked_count"].notna().any():
        row = _title(ws, row, "TOP 10 PRODUK PALING DISUKAI (LIKED)")
        data_start = row
        top_liked = df[df["liked_count"].notna()].nlargest(10, "liked_count")
        row = _header(ws, row, ["Nama Produk", "Harga", "Total Favorit", "Terjual", "Rating"])
        for _, r in top_liked.iterrows():
            row = _row(ws, row, [
                str(r.get("nama_produk", ""))[:45],
                str(r.get("harga", "")),
                int(r["liked_count"]) if pd.notna(r["liked_count"]) else "",
                int(r["terjual"]) if pd.notna(r.get("terjual")) else "",
                float(r["rating"]) if pd.notna(r.get("rating")) else "",
            ])
        _bar_chart(ws, "Top 10 Produk Paling Disukai", 3, 1, data_start, data_start + len(top_liked), data_start, horizontal=True)
        row = max(row, data_start + CHART_ROWS + 3) + SECTION_GAP

    # Section 5: Top 10 by Comment Count (banyak ulasan = banyak interaksi)
    if "comment_count" in df.columns and df["comment_count"].notna().any():
        row = _title(ws, row, "TOP 10 PRODUK PALING BANYAK ULASAN (COMMENT)")
        data_start = row
        top_cmc = df[df["comment_count"].notna()].nlargest(10, "comment_count")
        row = _header(ws, row, ["Nama Produk", "Harga", "Jumlah Ulasan", "Terjual", "Rating"])
        for _, r in top_cmc.iterrows():
            row = _row(ws, row, [
                str(r.get("nama_produk", ""))[:45],
                str(r.get("harga", "")),
                int(r["comment_count"]) if pd.notna(r["comment_count"]) else "",
                int(r["terjual"]) if pd.notna(r.get("terjual")) else "",
                float(r["rating"]) if pd.notna(r.get("rating")) else "",
            ])
        _bar_chart(ws, "Top 10 Produk Banyak Ulasan", 3, 1, data_start, data_start + len(top_cmc), data_start, horizontal=True)
        row = max(row, data_start + CHART_ROWS + 3) + SECTION_GAP

    # Section 6: Stock Analysis
    if "stock" in df.columns and df["stock"].notna().any():
        row = _title(ws, row, "ANALISIS STOCK")
        stock_data = df[df["stock"].notna()]["stock"]
        avg_stock = stock_data.mean()
        min_stock = stock_data.min()
        max_stock = stock_data.max()
        low_stock = int((stock_data < 10).sum())
        row = _header(ws, row, ["Metrik", "Nilai"])
        row = _row(ws, row, ["Rata-rata Stock", int(avg_stock)])
        row = _row(ws, row, ["Stock Tertinggi", int(max_stock)])
        row = _row(ws, row, ["Stock Terendah", int(min_stock)])
        row = _row(ws, row, ["Produk Stock < 10", low_stock])
        row += SECTION_GAP

    # Section 7: Interaksi Score (liked + comment combined)
    if "liked_count" in df.columns and "comment_count" in df.columns:
        df_inter = df.copy()
        df_inter["liked_safe"] = pd.to_numeric(df_inter.get("liked_count"), errors="coerce").fillna(0)
        df_inter["cmc_safe"] = pd.to_numeric(df_inter.get("comment_count"), errors="coerce").fillna(0)
        df_inter["interaksi_score"] = df_inter["liked_safe"] + df_inter["cmc_safe"]

        row = _title(ws, row, "TOP 10 PRODUK DENGAN INTERAKSI TERTINGGI")
        row = _subtitle(ws, row, "Skor = Liked Count + Comment Count")
        data_start = row
        top_inter = df_inter[df_inter["interaksi_score"] > 0].nlargest(10, "interaksi_score")
        row = _header(ws, row, ["Nama Produk", "Harga", "Total Favorit", "Jumlah Ulasan", "Interaksi Score"])
        for _, r in top_inter.iterrows():
            row = _row(ws, row, [
                str(r.get("nama_produk", ""))[:45],
                str(r.get("harga", "")),
                int(r["liked_safe"]) if r["liked_safe"] > 0 else "",
                int(r["cmc_safe"]) if r["cmc_safe"] > 0 else "",
                int(r["interaksi_score"]),
            ])
        _bar_chart(ws, "Top 10 Interaksi", 5, 1, data_start, data_start + len(top_inter), data_start, horizontal=True)
        row = max(row, data_start + CHART_ROWS + 3) + SECTION_GAP

    # Section 8: Seller Type vs Avg Price
    if "seller_type" in df.columns and "harga_angka" in df.columns:
        row = _title(ws, row, "RATA-RATA HARGA PER TIPE SELLER")
        avg_price_seller = df.groupby("seller_type")["harga_angka"].agg(["mean", "count"]).sort_values("mean", ascending=False)
        data_start = row
        row = _header(ws, row, ["Tipe Seller", "Rata-rata Harga", "Jumlah Produk"])
        for stype, r in avg_price_seller.iterrows():
            row = _row(ws, row, [str(stype), "Rp{:,.0f}".format(r["mean"]), int(r["count"])])
        _bar_chart(ws, "Harga Rata-rata per Tipe Seller", 2, 1, data_start, data_start + len(avg_price_seller), data_start, horizontal=True)

    _auto_width(ws, 5)


# ============================================================
# ANALISIS FIELD REKOMENDASI (AFFILIATOR) — SHEET v2: REKOMENDASI
# ============================================================

def _build_rekomendasi_v2(wb, df):
    """Sheet baru: Rekomendasi produk untuk affiliator/user."""
    ws = wb.create_sheet("Rekomendasi")
    row = 1

    # Score: kombinasi rating, liked, comments, terjual, diskon
    df_r = df.copy()
    df_r["rating_safe"] = pd.to_numeric(df_r.get("rating", pd.Series(dtype=float)), errors="coerce").fillna(0)
    df_r["terjual_safe"] = pd.to_numeric(df_r.get("terjual", pd.Series(dtype=float)), errors="coerce").fillna(0)
    df_r["harga_safe"] = df_r["harga_angka"].fillna(1)
    df_r["liked_safe"] = pd.to_numeric(df_r.get("liked_count", pd.Series(dtype=float)), errors="coerce").fillna(0)
    df_r["cmc_safe"] = pd.to_numeric(df_r.get("comment_count", pd.Series(dtype=float)), errors="coerce").fillna(0)
    df_r["diskon_safe"] = pd.to_numeric(df_r.get("diskon_persen", pd.Series(dtype=float)), errors="coerce").fillna(0)
    df_r["stock_safe"] = pd.to_numeric(df_r.get("stock", pd.Series(dtype=float)), errors="coerce").fillna(0)

    # Weight: terjual (paling penting), rating, liked, comment, diskon
    # Skor normalisasi 0-100
    def norm(series, higher=True):
        mn, mx = series.min(), series.max()
        if mx == mn:
            return pd.Series(50, index=series.index)
        normed = (series - mn) / (mx - mn) * 100
        return normed if higher else 100 - normed

    df_r["r_rating"] = norm(df_r["rating_safe"])
    df_r["r_terjual"] = norm(df_r["terjual_safe"])
    df_r["r_liked"] = norm(df_r["liked_safe"])
    df_r["r_cmc"] = norm(df_r["cmc_safe"])
    df_r["r_harga"] = norm(df_r["harga_safe"], higher=False)  # murah = lebih baik
    df_r["r_diskon"] = norm(df_r["diskon_safe"])
    df_r["r_stock"] = norm(df_r["stock_safe"])

    # Rekomendasi Score (weighted average)
    df_r["rekomendasi_score"] = round(
        df_r["r_rating"] * 0.15
        + df_r["r_terjual"] * 0.30
        + df_r["r_liked"] * 0.15
        + df_r["r_cmc"] * 0.10
        + df_r["r_harga"] * 0.15
        + df_r["r_diskon"] * 0.10
        + df_r["r_stock"] * 0.05,
        2
    )

    # === Section 1: Header Info ===
    ws.merge_cells("A1:J1")
    ws.cell(row=1, column=1,
            value="📋 REKOMENDASI PRODUK — BERDASARKAN ANALISIS MULTI-KRITERIA").font = Font(bold=True, size=13, color="1F4E79")
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
    row = 3

    ws.cell(row=row, column=1,
            value="Bobot: Terjual 30% | Rating 15% | Harga 15% | Liked 15% | Diskon 10% | Comment 10% | Stock 5%").font = SUBTITLE_FONT
    row += 1

    # === Section 2: Top Rekomendasi (Top 20) ===
    row = _title(ws, row, "TOP 20 REKOMENDASI PRODUK")
    top_rec = df_r[df_r["rekomendasi_score"] > 0].nlargest(20, "rekomendasi_score")
    data_start = row
    headers = ["Nama Produk", "Harga", "Rating", "Terjual", "Total Favorit", "Jumlah Ulasan",
               "Seller", "Diskon %", "Score"]
    row = _header(ws, row, headers)
    for _, r in top_rec.iterrows():
        row = _row(ws, row, [
            str(r.get("nama_produk", ""))[:50],
            str(r.get("harga", "")),
            float(r["rating_safe"]) if r["rating_safe"] > 0 else "",
            int(r["terjual_safe"]) if r["terjual_safe"] > 0 else "",
            int(r["liked_safe"]) if r["liked_safe"] > 0 else "",
            int(r["cmc_safe"]) if r["cmc_safe"] > 0 else "",
            str(r.get("penjual", ""))[:30],
            "{:.1f}%".format(r["diskon_safe"]) if r["diskon_safe"] > 0 else "-",
            float(r["rekomendasi_score"]),
        ])

    # Bar chart for scores
    _bar_chart(ws, "Skor Rekomendasi", 9, 1, data_start, data_start + len(top_rec), data_start, horizontal=True)
    row = max(row, data_start + CHART_ROWS + 3) + SECTION_GAP

    # === Section 3: Best Value Murah ===
    row = _title(ws, row, "BEST VALUE — HARGA DI BAWAH 5 JUTA")
    df_budget = df_r[(df_r["harga_angka"] < 5_000_000) & (df_r["harga_angka"] > 0)]
    budget_best = df_budget[df_budget["rekomendasi_score"] > 0].nlargest(10, "rekomendasi_score")
    row = _header(ws, row, ["Nama Produk", "Harga", "Rating", "Terjual", "Favorit", "Ulasan", "Seller", "Score"])
    for _, r in budget_best.iterrows():
        row = _row(ws, row, [
            str(r.get("nama_produk", ""))[:50], str(r.get("harga", "")),
            float(r["rating_safe"]) if r["rating_safe"] > 0 else "",
            int(r["terjual_safe"]) if r["terjual_safe"] > 0 else "",
            int(r["liked_safe"]) if r["liked_safe"] > 0 else "",
            int(r["cmc_safe"]) if r["cmc_safe"] > 0 else "",
            str(r.get("penjual", ""))[:30],
            float(r["rekomendasi_score"]),
        ])
    row += SECTION_GAP

    # === Section 4: Best Value Mid-Range ===
    row = _title(ws, row, "BEST VALUE — HARGA 5-15 JUTA")
    df_mid = df_r[(df_r["harga_angka"] >= 5_000_000) & (df_r["harga_angka"] < 15_000_000)]
    mid_best = df_mid[df_mid["rekomendasi_score"] > 0].nlargest(10, "rekomendasi_score")
    row = _header(ws, row, ["Nama Produk", "Harga", "Rating", "Terjual", "Favorit", "Ulasan", "Seller", "Score"])
    for _, r in mid_best.iterrows():
        row = _row(ws, row, [
            str(r.get("nama_produk", ""))[:50], str(r.get("harga", "")),
            float(r["rating_safe"]) if r["rating_safe"] > 0 else "",
            int(r["terjual_safe"]) if r["terjual_safe"] > 0 else "",
            int(r["liked_safe"]) if r["liked_safe"] > 0 else "",
            int(r["cmc_safe"]) if r["cmc_safe"] > 0 else "",
            str(r.get("penjual", ""))[:30],
            float(r["rekomendasi_score"]),
        ])
    row += SECTION_GAP

    # === Section 5: Best Value Premium ===
    row = _title(ws, row, "BEST VALUE — HARGA DI ATAS 15 JUTA")
    df_prem = df_r[df_r["harga_angka"] >= 15_000_000]
    prem_best = df_prem[df_prem["rekomendasi_score"] > 0].nlargest(10, "rekomendasi_score")
    row = _header(ws, row, ["Nama Produk", "Harga", "Rating", "Terjual", "Favorit", "Ulasan", "Seller", "Score"])
    for _, r in prem_best.iterrows():
        row = _row(ws, row, [
            str(r.get("nama_produk", ""))[:50], str(r.get("harga", "")),
            float(r["rating_safe"]) if r["rating_safe"] > 0 else "",
            int(r["terjual_safe"]) if r["terjual_safe"] > 0 else "",
            int(r["liked_safe"]) if r["liked_safe"] > 0 else "",
            int(r["cmc_safe"]) if r["cmc_safe"] > 0 else "",
            str(r.get("penjual", ""))[:30],
            float(r["rekomendasi_score"]),
        ])
    row += SECTION_GAP

    # === Section 6: High Engagement (banyak liked + comment) ===
    if "liked_safe" in df_r.columns and "cmc_safe" in df_r.columns:
        row = _title(ws, row, "PRODUK DENGAN ENGAGEMENT TERTINGGI")
        row = _subtitle(ws, row, "Produk yang paling banyak mendapat perhatian (favorit + ulasan)")
        df_r["engagement_score"] = df_r["liked_safe"] + df_r["cmc_safe"]
        top_eng = df_r[df_r["engagement_score"] > 0].nlargest(10, "engagement_score")
        row = _header(ws, row, ["Nama Produk", "Total Favorit", "Jumlah Ulasan", "Harga", "Seller", "Engagement"])
        for _, r in top_eng.iterrows():
            row = _row(ws, row, [
                str(r.get("nama_produk", ""))[:50],
                int(r["liked_safe"]) if r["liked_safe"] > 0 else "",
                int(r["cmc_safe"]) if r["cmc_safe"] > 0 else "",
                str(r.get("harga", "")),
                str(r.get("penjual", ""))[:30],
                int(r["engagement_score"]),
            ])
        row += SECTION_GAP

    # === Section 7: Official / Flagship Seller ===
    if "seller_type" in df_r.columns:
        row = _title(ws, row, "REKOMENDASI PRODUK DARI SELLER OFFICIAL/FLAGSHIP")
        official = df_r[df_r["seller_type"].notna()
                       & df_r["seller_type"].str.contains("Official|Flagship|Mall", case=False, na=False)]
        official_scored = official[official["rekomendasi_score"] > 0].nlargest(10, "rekomendasi_score")
        row = _header(ws, row, ["Nama Produk", "Harga", "Rating", "Terjual", "Favorit", "Ulasan", "Tipe Seller", "Score"])
        for _, r in official_scored.iterrows():
            row = _row(ws, row, [
                str(r.get("nama_produk", ""))[:50], str(r.get("harga", "")),
                float(r["rating_safe"]) if r["rating_safe"] > 0 else "",
                int(r["terjual_safe"]) if r["terjual_safe"] > 0 else "",
                int(r["liked_safe"]) if r["liked_safe"] > 0 else "",
                int(r["cmc_safe"]) if r["cmc_safe"] > 0 else "",
                str(r.get("seller_type", "")),
                float(r["rekomendasi_score"]),
            ])
        row += SECTION_GAP

    # === Section 8: Flash Sale Rekomendasi ===
    if "flash_sale" in df_r.columns and df_r["flash_sale"].notna().any():
        row = _title(ws, row, "REKOMENDASI PRODUK FLASH SALE")
        flash_rec = df_r[df_r["flash_sale"] == "Ya"]
        if len(flash_rec) > 0:
            flash_best = flash_rec[flash_rec["rekomendasi_score"] > 0].nlargest(10, "rekomendasi_score")
            row = _header(ws, row, ["Nama Produk", "Harga", "Diskon %", "Terjual", "Favorit", "Ulasan", "Seller", "Score"])
            for _, r in flash_best.iterrows():
                row = _row(ws, row, [
                    str(r.get("nama_produk", ""))[:50], str(r.get("harga", "")),
                    "{:.1f}%".format(r["diskon_safe"]) if r["diskon_safe"] > 0 else "-",
                    int(r["terjual_safe"]) if r["terjual_safe"] > 0 else "",
                    int(r["liked_safe"]) if r["liked_safe"] > 0 else "",
                    int(r["cmc_safe"]) if r["cmc_safe"] > 0 else "",
                    str(r.get("penjual", ""))[:30],
                    float(r["rekomendasi_score"]),
                ])
        else:
            row = _row(ws, row, ["Tidak ada produk flash sale"])
        row += SECTION_GAP

    # === Section 9: Score Breakdown ===
    row = _title(ws, row, "DETAIL SKOR — SEMUA PRODUK (TOP 50)")
    ws.cell(row=row, column=1,
            value="Terjual, Rating, Harga, Liked, Diskon, Comment, Stock masing-masing dinormalisasi 0-100").font = SUBTITLE_FONT
    row += 1
    detail = df_r[df_r["rekomendasi_score"] > 0].nlargest(50, "rekomendasi_score")
    row = _header(ws, row, ["Nama Produk", "Harga", "rTerjual", "rRating", "rHarga", "rLiked", "rDiskon", "rComment", "Score"])
    for _, r in detail.iterrows():
        row = _row(ws, row, [
            str(r.get("nama_produk", ""))[:40],
            str(r.get("harga", "")),
            round(float(r["r_terjual"]), 1),
            round(float(r["r_rating"]), 1),
            round(float(r["r_harga"]), 1),
            round(float(r["r_liked"]), 1),
            round(float(r["r_diskon"]), 1),
            round(float(r["r_cmc"]), 1),
            float(r["rekomendasi_score"]),
        ])

    _auto_width(ws, 9)




def _build_dashboard(wb, df, keyword, filters=None, marketplace="blibli"):
    ws = wb.create_sheet("Ecommerce Scraping")
    harga = df["harga_angka"].dropna()
    total = len(df)

    row = 1
    ws.merge_cells("A1:D1")
    mp_title = marketplace.upper()
    ws.cell(row=1, column=1, value="{} SCRAPING - ECOMMERCE ANALYTICS".format(mp_title)).font = Font(bold=True, size=14, color="1F4E79")
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
    row = 3

    # KEY METRICS
    row = _title(ws, row, "KEY METRICS")
    row = _header(ws, row, ["Metrik", "Nilai"])
    metrics = [
        ["Total Produk", total],
        ["Harga Terendah", "Rp{:,.0f}".format(harga.min()) if len(harga) else "N/A"],
        ["Harga Tertinggi", "Rp{:,.0f}".format(harga.max()) if len(harga) else "N/A"],
        ["Harga Rata-rata", "Rp{:,.0f}".format(harga.mean()) if len(harga) else "N/A"],
        ["Harga Median", "Rp{:,.0f}".format(harga.median()) if len(harga) else "N/A"],
        ["Jumlah Seller Unik", int(df["penjual"].nunique()) if "penjual" in df.columns else "N/A"],
        ["Jumlah Kota", int(df["kota"].nunique()) if "kota" in df.columns else "N/A"],
        ["Rata-rata Rating", "{:.2f}".format(df["rating"].mean()) if "rating" in df.columns and df["rating"].notna().any() else "N/A"],
        ["Total Terjual", "{:,}".format(int(df["terjual"].sum())) if "terjual" in df.columns and df["terjual"].notna().any() else "N/A"],
        ["Produk Diskon", int(df["diskon_persen"].notna().sum())],
        ["Avg Diskon", "{:.1f}%".format(df["diskon_persen"].mean()) if df["diskon_persen"].notna().any() else "N/A"],
        # --- Field Rekomendasi ---
        ["Total SKU", int(df["item_id"].notna().sum()) if "item_id" in df.columns else "N/A"],
        ["Produk Ada Stock", int(df["stock"].notna().sum()) if "stock" in df.columns else "N/A"],
        ["Total Favorit", "{:,}".format(int(df["liked_count"].sum())) if "liked_count" in df.columns and df["liked_count"].notna().any() else "N/A"],
        ["Total Jumlah Ulasan", "{:,}".format(int(df["comment_count"].sum())) if "comment_count" in df.columns and df["comment_count"].notna().any() else "N/A"],
        ["Free Shipping", int(df["free_shipping"].notna().sum()) if "free_shipping" in df.columns else "N/A"],
        ["Flash Sale", int(df["flash_sale"].notna().sum()) if "flash_sale" in df.columns else "N/A"],
    ]
    for label, val in metrics:
        row = _row(ws, row, [label, val])

    row += SECTION_GAP

    # FILTERS
    row = _title(ws, row, "FILTER YANG DIGUNAKAN")
    row = _header(ws, row, ["Filter", "Nilai"])
    row = _row(ws, row, ["Keyword", keyword])
    row = _row(ws, row, ["Waktu Scraping", datetime.now().strftime("%Y-%m-%d %H:%M")])
    if filters:
        if filters.get("min_price") or filters.get("max_price"):
            row = _row(ws, row, ["Rentang Harga", "Rp{:,} - Rp{:,}".format(
                filters.get("min_price", 0), filters.get("max_price", 999999999))])
        if filters.get("min_rating"):
            row = _row(ws, row, ["Rating Minimum", str(filters["min_rating"]) + "+"])
        if filters.get("location"):
            row = _row(ws, row, ["Lokasi", filters["location"]])
        if filters.get("sort", 0) != 0:
            sorts = {1: "Terbaru", 2: "Termurah", 3: "Termahal", 4: "Terlaris"}
            row = _row(ws, row, ["Urutan", sorts.get(filters["sort"], "Default")])

    row += SECTION_GAP

    # DATA COMPLETENESS
    row = _title(ws, row, "KELENGKAPAN DATA")
    row = _header(ws, row, ["Field", "Terisi", "Kosong", "Rate"])
    for field in ["nama_produk", "harga", "penjual", "kota", "terjual", "rating",
                  # Field Rekomendasi
                  "item_id", "stock", "liked_count", "comment_count", "free_shipping", "seller_type"]:
        if field in df.columns:
            friendly_name = {"item_id": "SKU", "liked_count": "Total Favorit", "comment_count": "Jumlah Ulasan"}.get(field, field)
            complete = int(df[field].notna().sum())
            if df[field].dtype == "object":
                complete = int((df[field].notna() & (df[field] != "")).sum())
            missing = total - complete
            rate = "{:.0f}%".format(complete / total * 100) if total else "0%"
            row = _row(ws, row, [friendly_name, complete, missing, rate])

    row += SECTION_GAP

    # TOP 5 SELLERS
    row = _title(ws, row, "TOP 5 SELLER")
    row = _header(ws, row, ["Seller", "Produk", "Avg Harga"])
    if "penjual" in df.columns:
        for seller, grp in df.groupby("penjual").size().sort_values(ascending=False).head(5).items():
            avg_h = df[df["penjual"] == seller]["harga_angka"].mean()
            row = _row(ws, row, [str(seller)[:35], int(grp), "Rp{:,.0f}".format(avg_h) if pd.notna(avg_h) else ""])

    row += SECTION_GAP

    # TOP 5 TERLARIS
    row = _title(ws, row, "TOP 5 PRODUK TERLARIS")
    if "terjual" in df.columns and df["terjual"].notna().any():
        row = _header(ws, row, ["Produk", "Harga", "Terjual"])
        for _, r in df[df["terjual"].notna()].nlargest(5, "terjual").iterrows():
            row = _row(ws, row, [str(r["nama_produk"])[:45], str(r.get("harga", "")), int(r["terjual"])])

    row += SECTION_GAP

    # TOP 5 DISKON
    row = _title(ws, row, "TOP 5 DISKON TERBESAR")
    if df["diskon_persen"].notna().any():
        row = _header(ws, row, ["Produk", "Harga", "Diskon"])
        for _, r in df[df["diskon_persen"].notna()].nlargest(5, "diskon_persen").iterrows():
            row = _row(ws, row, [str(r["nama_produk"])[:45], str(r.get("harga", "")), "{:.1f}%".format(r["diskon_persen"])])

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 12
